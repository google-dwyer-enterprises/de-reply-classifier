"""Build the 'ecom_industries' sheet in prospeo_industry_candidates.xlsx.

Reads 'all_industries', applies an e-com keyword filter, then dedupes by
normalized form (strip leading slash, lowercase, ' & ' -> ' and ',
collapse whitespace). Keeps the highest-lead_count variant as the display
name; sums counts across variants."""

import re
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent / "exports" / "prospeo_industry_candidates.xlsx"

INCLUDE = re.compile(
    r"\b("
    r"apparel|fashion|clothing|shoes?|footwear|jewelry|jewellery|accessor\w*|"
    r"beauty|cosmetic\w*|personal care|skin|skincare|hair care|haircare|make[- ]?up|fragrance|perfume|"
    r"home|kitchen|furniture|decor|garden|bedding|bath|"
    r"food|beverage|grocer\w*|gourmet|snack\w*|nutrition|supplement\w*|coffee|tea|wine|spirit\w*|"
    r"household|wellness|"
    r"sport\w*|outdoor\w*|fitness|athletic|gym|"
    r"pet\b|pets\b|pet supplies|pet care|"
    r"toys?|games?|baby|infant|kids?|child\w*|"
    r"electronic\w*|gadget\w*|"
    r"consumer goods|consumer product\w*|consumer services|"
    r"e[- ]?commerce|"
    r"book\w*|"
    r"automotive parts?|auto parts?"
    r")\b",
    re.IGNORECASE,
)

EXCLUDE = re.compile(
    r"\b("
    r"software|saas|it services|information technology|cyber|cloud|"
    r"marketing services|advertising|agency|"
    r"finance|insurance|bank|investment|"
    r"hospital|clinic|medical practice|pharmaceutical|biotech|"
    r"education|university|school|"
    r"real estate|construction|architectural|civil engineer|"
    r"law|legal|"
    r"oil|gas|mining|chemical|"
    r"hospitality|hotel|restaurant|"
    r"telecommunication|telecom|broadcast|"
    r"defense|military|aerospace|aviation|"
    r"logistics|shipping|trucking|warehouse|"
    r"staffing|recruiting|human resources|consulting"
    r")\b",
    re.IGNORECASE,
)


def normalize(name: str) -> str:
    s = name.strip().lstrip("/").lower()
    s = s.replace(" & ", " and ")
    s = re.sub(r"\s+", " ", s)
    return s


wb = load_workbook(XLSX)
ws_all = wb["all_industries"]
rows = list(ws_all.iter_rows(values_only=True))
headers = [str(h) if h is not None else "" for h in rows[0]]
i_ind = headers.index("industry")
i_n = headers.index("lead_count")
i_compat = headers.index("prospeo_compatible")
i_sugg = headers.index("suggested_prospeo_value")

# Group by normalized form. Slash-paths are split into individual segments so
# each becomes a Prospeo-probeable string. One lead may contribute to multiple
# segments — lead_count below is "leads whose original tag contained this
# segment", NOT unique-lead totals. Don't sum across rows.
groups: dict[str, list[tuple]] = defaultdict(list)
for r in rows[1:]:
    if not r or not r[i_ind]:
        continue
    name = str(r[i_ind])
    if not INCLUDE.search(name) or EXCLUDE.search(name):
        continue
    count = r[i_n] or 0

    if name.startswith("/"):
        # Path — explode into segments, each becomes its own candidate row
        segments = [s.strip() for s in name.lstrip("/").split("/") if s.strip()]
        for seg in segments:
            # Apply include/exclude per segment too (e.g. drop generic "Other")
            if not INCLUDE.search(seg) or EXCLUDE.search(seg):
                continue
            key = normalize(seg)
            groups[key].append((seg, count, r[i_compat], r[i_sugg]))
    else:
        key = normalize(name)
        groups[key].append((name, count, r[i_compat], r[i_sugg]))

# Collapse: pick highest-count variant as display, sum counts
deduped = []
for key, variants in groups.items():
    variants.sort(key=lambda v: -v[1])
    display, _, compat, sugg = variants[0]
    total = sum(v[1] for v in variants)
    deduped.append((display, total, compat, sugg, len(variants)))

deduped.sort(key=lambda x: (-x[1], x[0]))

# Write
if "ecom_industries" in wb.sheetnames:
    del wb["ecom_industries"]
ws = wb.create_sheet("ecom_industries")
ws.append(["industry", "lead_count", "prospeo_compatible",
           "suggested_prospeo_value", "variant_count"])
for d in deduped:
    ws.append(list(d))

for col_idx, name in enumerate(["industry", "lead_count", "prospeo_compatible",
                                 "suggested_prospeo_value", "variant_count"], start=1):
    max_len = max([len(name)] + [len(str(d[col_idx - 1])) for d in deduped] or [0])
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

wb.save(XLSX)
total_variants = sum(d[4] for d in deduped)
print(f"Pre-dedupe rows kept by filter: {total_variants}")
print(f"Post-dedupe distinct industries: {len(deduped)}")
print(f"Top 20 (deduped, by total lead_count):")
for d in deduped[:20]:
    print(f"  {d[1]:>6}  variants={d[4]:>3}  {d[0]}")
