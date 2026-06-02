"""Final round: discover which e-com industry strings are actually accepted.
Each rejection is free (400 with filter_error tells us what's invalid).
We only spend credits on calls that return 200.

Two modes:
  - No args: probe the hardcoded e-com candidate list (original behavior).
  - --from-xlsx <path>: probe every distinct industry string in the xlsx (both
    sheets) whose prospeo_compatible == 'unknown'. Appends a probe_results
    sheet to the same workbook so verdicts can be promoted on the next run.

Flags:
  --dry-run   Print what would be probed + worst-case credit cost. No API calls.
  --yes       Skip the y/N confirmation prompt.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import requests
from dotenv import load_dotenv

from prospeo_sync import PROSPEO_BASE


# Hand-crafted candidates aimed at the 5 unverified set-1 gaps:
# supplements, alt medicine, health/wellness, mechanical goods, electronics.
# Mostly "Retail X" and "X Manufacturing" — patterns that worked for the
# verified-10 (Retail Apparel and Fashion, Retail Groceries, Apparel
# Manufacturing, etc.).
GAP_CANDIDATES = [
    # Supplements / nutrition
    "Dietary Supplement Manufacturing",
    "Nutritional Supplement Manufacturing",
    "Vitamins and Dietary Supplements Manufacturing",
    "Pharmaceutical Manufacturing",
    # Alt medicine
    "Alternative Medicine",
    "Holistic Health Services",
    "Naturopathic Medicine",
    # Health & wellness
    "Wellness and Fitness Services",
    "Health and Human Services",
    "Retail Health and Personal Care Products",
    # Mechanical goods / tools
    "Machinery Manufacturing",
    "Tool and Hardware Manufacturing",
    "Hardware Manufacturing",
    "Industrial Machinery Manufacturing",
    # Electronics
    "Consumer Electronics Manufacturing",
    "Computers and Electronics Manufacturing",
    "Appliances, Electrical, and Electronics Manufacturing",
    "Electrical Equipment Manufacturing",
    "Retail Appliances, Electrical, and Electronic Equipment",
]


HARDCODED_CANDIDATES = [
    # Apparel-ish
    "Retail Apparel and Fashion",
    "Apparel Manufacturing",
    "Apparel & Fashion",
    "Fashion",
    "Apparel",
    "Retail",
    # Beauty / personal care
    "Cosmetics",
    "Personal Care Product Manufacturing",
    "Personal Care Products",
    "Health and Beauty",
    # Food / beverage
    "Food and Beverage Services",
    "Food and Beverage Manufacturing",
    "Food Production",
    "Beverage Manufacturing",
    # Home / furniture
    "Furniture and Home Furnishings Manufacturing",
    "Retail Home Furnishings",
    "Furniture",
    # Sports / outdoors
    "Sporting Goods Manufacturing",
    "Spectator Sports",
    # Consumer goods catchall
    "Consumer Goods",
    "Consumer Services",
    "Retail Consumer Goods",
    # Pet
    "Pet Services",
    # Health / wellness
    "Wellness and Fitness Services",
    "Health, Wellness & Fitness",
    # Generic manufacturing
    "Manufacturing",
    # Known-working sanity check
    "Software Development",
]


def call(filters: dict, api_key: str) -> tuple[int, dict]:
    url = f"{PROSPEO_BASE}/search-person"
    r = requests.post(url, json={"filters": filters, "page": 1},
                       headers={"X-KEY": api_key, "Content-Type": "application/json"},
                       timeout=20)
    try:
        body = r.json()
    except ValueError:
        body = {}
    return r.status_code, body


def probe_industry(value: str, api_key: str) -> tuple[str, int, int | None, str]:
    """Returns (verdict, http_status, total_count_or_None, error_or_message)."""
    status, body = call({"company_industry": {"include": [value]}}, api_key)
    if status == 200:
        tc = (body.get("pagination") or {}).get("total_count")
        return ("accepted", status, tc, "")
    if status == 400:
        fe = body.get("filter_error") or body.get("error_toast") or ""
        return ("rejected", status, None, str(fe))
    return ("error", status, None, json.dumps(body)[:200])


def probe_country(countries: list[str], api_key: str) -> tuple[str, int, int | None, str]:
    """Probe company_location_search with a list of country strings.

    Prospeo's docs say location values must be Suggestions-API-resolved.
    Raw strings may be rejected as INVALID_FILTERS, silently ignored, or
    (best case) accepted. Returns same shape as probe_industry."""
    status, body = call({"company_location_search": {"include": countries}}, api_key)
    if status == 200:
        tc = (body.get("pagination") or {}).get("total_count")
        return ("accepted", status, tc, "")
    if status == 400:
        fe = body.get("filter_error") or body.get("error_toast") or ""
        return ("rejected", status, None, str(fe))
    return ("error", status, None, json.dumps(body)[:200])


def load_unknowns_from_xlsx(path: Path, sheets: list[str] | None = None) -> list[str]:
    """Read 'industry' columns where prospeo_compatible == 'unknown'.

    sheets: which tabs to read. None means both ('all_industries',
    'positive_industries'); pass a single-element list to restrict.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    unknowns: set[str] = set()
    target_sheets = sheets if sheets else ["all_industries", "positive_industries"]
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"  (sheet {sheet_name!r} not found in workbook; skipping)")
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(rows)]
        try:
            i_ind = headers.index("industry")
            i_compat = headers.index("prospeo_compatible")
        except ValueError:
            print(f"  (sheet {sheet_name!r} missing expected columns; skipping)")
            continue
        for row in rows:
            if not row:
                continue
            ind = row[i_ind]
            compat = row[i_compat]
            if compat == "unknown" and ind and str(ind).strip():
                unknowns.add(str(ind))
    return sorted(unknowns)


def write_probe_results(path: Path, results: list[dict]) -> None:
    """Append/replace a 'probe_results' sheet on the workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path))
    if "probe_results" in wb.sheetnames:
        del wb["probe_results"]
    ws = wb.create_sheet("probe_results")
    ws.append(["industry", "verdict", "http_status", "total_count", "message"])
    for r in results:
        ws.append([r["industry"], r["verdict"], r["http_status"],
                   r["total_count"], r["message"]])
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--gaps", action="store_true",
                        help="Probe the hand-crafted set-1 gap candidates "
                             "(supplements, alt medicine, wellness, mechanical, "
                             "electronics). Mutually exclusive with --from-xlsx.")
    parser.add_argument("--probe-country", action="store_true",
                        help="Probe company_location_search with United States "
                             "and Canada — individually, then combined. "
                             "Worst case 3 credits. Mutually exclusive with "
                             "--from-xlsx and --gaps.")
    parser.add_argument("--from-xlsx", type=Path, default=None,
                        help="Read unknown industries from this xlsx.")
    parser.add_argument("--sheet", choices=["all", "positive", "ecom", "both"],
                        default="both",
                        help="Which tab to read when --from-xlsx is set "
                             "(all=all_industries, positive=positive_industries, "
                             "ecom=ecom_industries, both=all+positive union; "
                             "default: both).")
    parser.add_argument("--dry-run", action="store_true",
                        help="List what would be probed. No API calls.")
    parser.add_argument("--yes", action="store_true",
                        help="Skip confirmation prompt.")
    args = parser.parse_args()

    load_dotenv()
    api_key = os.environ.get("PROSPEO_API_KEY", "").strip()

    mutually_exclusive = sum(bool(x) for x in
                              (args.gaps, args.from_xlsx, args.probe_country))
    if mutually_exclusive > 1:
        sys.exit("Use only one of --gaps / --from-xlsx / --probe-country.")

    if args.probe_country:
        # Special-case: probe location filter (different shape, different
        # endpoint behavior). Bypass the candidate-list machinery below.
        if args.dry_run:
            print("DRY RUN — would probe company_location_search with:")
            print("  1. {'include': ['United States']}")
            print("  2. {'include': ['Canada']}")
            print("  3. {'include': ['United States', 'Canada']}")
            print("Worst case 3 credits (each rejection is free).")
            return
        if not api_key:
            sys.exit("PROSPEO_API_KEY not set in .env")
        if not args.yes:
            resp = input("Probe location filter (worst case 3 credits)? [y/N] ").strip().lower()
            if resp != "y":
                print("Aborted.")
                return
        for label, countries in [
            ("United States only", ["United States"]),
            ("Canada only", ["Canada"]),
            ("United States + Canada", ["United States", "Canada"]),
        ]:
            verdict, status, tc, msg = probe_country(countries, api_key)
            if verdict == "accepted":
                tail = f"OK (HTTP 200, total_count={tc}) — credit spent"
            elif verdict == "rejected":
                tail = f"REJECTED — {msg}"
            else:
                tail = f"HTTP {status}: {msg}"
            print(f"  {label:30s} {tail}")
        print("\nInterpretation:")
        print("  - 200 + small total_count → raw strings work, filter is real.")
        print("  - 400 INVALID_FILTERS    → values need Suggestions-API resolution.")
        print("  - 200 + huge total_count → silently ignored (compare against")
        print("    a no-location baseline; should be tens of millions if so).")
        return

    if args.gaps:
        candidates = list(GAP_CANDIDATES)
        print(f"Using {len(candidates)} hand-crafted gap candidates "
              "(supplements / alt medicine / wellness / mechanical / electronics).")
    elif args.from_xlsx:
        if not args.from_xlsx.exists():
            sys.exit(f"File not found: {args.from_xlsx}")
        sheet_map = {
            "all": ["all_industries"],
            "positive": ["positive_industries"],
            "ecom": ["ecom_industries"],
            "both": ["all_industries", "positive_industries"],
        }
        candidates = load_unknowns_from_xlsx(args.from_xlsx, sheet_map[args.sheet])
        print(f"Loaded {len(candidates)} 'unknown' industries from "
              f"{args.from_xlsx} (sheet={args.sheet})")
    else:
        candidates = HARDCODED_CANDIDATES
        print(f"Using {len(candidates)} hardcoded candidates.")

    if not candidates:
        print("Nothing to probe.")
        return

    print(f"\nWorst-case credit cost: {len(candidates)} (1 per HTTP 200; "
          "rejections are free).")

    if args.dry_run:
        print("\nDRY RUN — values that would be probed:")
        for v in candidates:
            print(f"  - {v}")
        return

    if not api_key:
        sys.exit("PROSPEO_API_KEY not set in .env")

    if not args.yes:
        resp = input("Proceed? [y/N] ").strip().lower()
        if resp != "y":
            print("Aborted.")
            return

    print("\nProbing industry enum values (each rejection is FREE):")
    results: list[dict] = []
    for v in candidates:
        verdict, http_status, tc, msg = probe_industry(v, api_key)
        results.append({
            "industry": v,
            "verdict": verdict,
            "http_status": http_status,
            "total_count": tc,
            "message": msg,
        })
        if verdict == "accepted":
            tail = f"OK (HTTP 200, total_count={tc}) — credit spent"
        elif verdict == "rejected":
            tail = f"REJECTED — {msg}"
        else:
            tail = f"HTTP {http_status}: {msg}"
        print(f"  {v!r:55s} {tail}")

    accepted = [r["industry"] for r in results if r["verdict"] == "accepted"]
    rejected = [r["industry"] for r in results if r["verdict"] == "rejected"]
    errored = [r["industry"] for r in results if r["verdict"] == "error"]

    print(f"\n=== ACCEPTED ({len(accepted)}): ===")
    for v in accepted:
        print(f"  - {v}")
    print(f"\n=== REJECTED ({len(rejected)}): ===")
    for v in rejected:
        print(f"  - {v}")
    if errored:
        print(f"\n=== ERROR ({len(errored)}): ===")
        for v in errored:
            print(f"  - {v}")

    if args.from_xlsx:
        write_probe_results(args.from_xlsx, results)
        print(f"\nWrote probe_results sheet -> {args.from_xlsx}")


if __name__ == "__main__":
    main()
