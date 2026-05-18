"""Pilot: compare Prospeo CATEGORY-mode vs DOMAIN-mode search-person.

Question this answers (Victor's feedback):
  "If we use Prospeo's industry/category filters instead of our domain list,
   can we get to 5,000 leads/week — and is the quality acceptable after the
   agency-filter cleanup?"

How it works:
  1. Pass A (DOMAIN mode)   — pick N random domains from inclusion_clean.csv,
                              run search-person + enrich-person + agency_filter.
  2. Pass B (CATEGORY mode) — run search-person with a curated list of e-com
                              industries (no domain filter), same downstream
                              pipeline.
  3. Both passes log every LLM accept/reject decision with the model's
     reasoning. Output is one XLSX with two sheets (one per pass) plus a
     summary line for direct comparison.

Hard budget cap via --max-credits (default 20 = ~$0.40). Pause-and-confirm
prompt before any live API call. --dry-run mode skips API entirely.

Run:
  python scripts/prospeo_category_pilot.py --dry-run            # no API, no spend
  python scripts/prospeo_category_pilot.py --max-credits 20     # live, capped

The script does NOT write to the database. Output goes to exports/category_pilot_<timestamp>.xlsx.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import anthropic
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from prospeo_sync import (
    _norm_domain, AGENCY_TOKENS, MARKETPLACE_DOMAINS, AGENCY_FILTER_MODEL,
    AGENCY_FILTER_PROMPT, PROSPEO_TIMEOUT, PROSPEO_BASE, PROSPEO_INDUSTRIES,
)


# Industry list is now the single source of truth in prospeo_sync.py.
# Keep PILOT_INDUSTRIES as an alias so older references continue to work.
PILOT_INDUSTRIES = PROSPEO_INDUSTRIES

# Mirrors PROSPEO_TITLES in prospeo_sync.py (two-tier: owner + marketing/e-com).
PILOT_TITLES = [
    # Tier 1: owner (15.8% lead-level conversion)
    "CEO", "Chief Executive Officer", "Founder", "Co-Founder", "Owner",
    "Co-Owner", "President", "Founder and CEO", "CEO and Founder",
    "Managing Director", "Chairman",
    # Tier 2: marketing/e-com (8.6% lead-level conversion)
    "CMO", "Chief Marketing Officer", "Head of Marketing", "VP Marketing",
    "Head of E-commerce", "Director of E-commerce",
]

DOMAIN_SAMPLE_SIZE = 60   # Oversample so we still hit 30 enriched results
CATEGORY_SAMPLE_SIZE = 30
DEFAULT_MAX_CREDITS = 40  # Worst case: 2 search pages/pass + 30 enriches/pass ~= 30-36

# Prospeo silently ignores the websites filter above 15 entries (verified in
# production after a 506-credit burn). NEVER send more than this per batch.
PROSPEO_BATCH_DOMAINS = 15
# If a websites-filtered batch returns more than this, the filter was silently
# ignored — abort the batch to prevent runaway credit spend.
PROSPEO_MAX_RESULTS_PER_BATCH = PROSPEO_BATCH_DOMAINS * 5

INCLUSION_CSV = Path(__file__).resolve().parent.parent / "original_data" / "inclusion_clean.csv"


# ---------- Prospeo HTTP ----------

def _search_person(filters: dict, api_key: str, page: int = 1) -> tuple[dict, int]:
    """One call to /search-person with the given filters dict.
    Returns (response_body, credits_used). Treats 400/NO_RESULTS as zero hits."""
    url = f"{PROSPEO_BASE}/search-person"
    headers = {"X-KEY": api_key, "Content-Type": "application/json"}
    body = {"filters": filters, "page": page}
    resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    if resp.status_code == 429:
        time.sleep(5)
        resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    if resp.status_code == 400:
        try:
            payload = resp.json()
        except ValueError:
            payload = {}
        if payload.get("error_code") == "NO_RESULTS":
            return ({"results": [], "pagination": {"total_count": 0, "total_page": 0}}, 0)
        resp.raise_for_status()
    resp.raise_for_status()
    return (resp.json() or {}, 1)


def _enrich_person(person_id: str, api_key: str) -> tuple[dict, int]:
    """Returns (enriched_dict, credits_used). 1 credit only if a verified email
    was returned; 0 otherwise. 400/NO_MATCH is a soft skip (no email available)."""
    url = f"{PROSPEO_BASE}/enrich-person"
    headers = {"X-KEY": api_key, "Content-Type": "application/json"}
    body = {"data": {"person_id": person_id}, "only_verified_email": True}
    resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    if resp.status_code == 400:
        return ({}, 0)
    if resp.status_code == 429:
        time.sleep(5)
        resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    resp.raise_for_status()
    data = resp.json() or {}
    if data.get("error"):
        return ({}, 0)
    person = data.get("person") or {}
    company = data.get("company") or {}
    email_block = person.get("email") or {}
    email = (email_block.get("email")
             if email_block.get("status") != "UNAVAILABLE" else None)
    billed = 0 if data.get("free_enrichment") else (1 if email else 0)
    return ({
        "email": email,
        "first_name": person.get("first_name"),
        "last_name": person.get("last_name"),
        "title": person.get("current_job_title"),
        "company_name": company.get("name") or company.get("company_name"),
        "company_domain": _norm_domain(company.get("website") or company.get("url")),
        "company_website": company.get("website") or company.get("url"),
        "company_description": company.get("description"),
    }, billed)


# ---------- Agency filter (mirrors prospeo_sync.py logic) ----------

def _email_domain(email: str | None) -> str | None:
    if not email or "@" not in email:
        return None
    return email.rsplit("@", 1)[1].lower()


def rule_classify(lead: dict) -> tuple[str, str, str] | None:
    name = (lead.get("company_name") or "").lower()
    site = _norm_domain(lead.get("company_website")) or ""
    edom = _email_domain(lead.get("email")) or ""
    if site in MARKETPLACE_DOMAINS or edom in MARKETPLACE_DOMAINS:
        return ("marketplace", "rule", "known marketplace domain")
    for tok in AGENCY_TOKENS:
        if tok in name or tok in site:
            return ("agency", "rule", f"matched agency token '{tok}'")
    return None


def llm_classify_one(client: anthropic.Anthropic, system_prompt: str,
                     lead: dict) -> tuple[str, str]:
    payload = {
        "company_name": lead.get("company_name"),
        "company_website": lead.get("company_website"),
        "company_description": lead.get("company_description"),
        "title": lead.get("title"),
        "email_domain": _email_domain(lead.get("email")),
    }
    resp = client.messages.create(
        model=AGENCY_FILTER_MODEL,
        max_tokens=200,
        system=system_prompt,
        messages=[{"role": "user", "content": json.dumps(payload)}],
    )
    text = resp.content[0].text.strip() if resp.content else "{}"
    if text.startswith("```"):
        text = text.split("```", 2)[1]
        if text.startswith("json"):
            text = text[4:]
        text = text.strip().rstrip("`").strip()
    try:
        parsed = json.loads(text)
        return (parsed.get("result", "unknown"), parsed.get("reason", "")[:200])
    except json.JSONDecodeError:
        return ("unknown", f"unparseable: {text[:80]}")


# ---------- Pass runners ----------

def _build_domain_filter(domains: list[str]) -> dict:
    return {
        "company": {"websites": {"include": domains}},
        "person_job_title": {
            "include": PILOT_TITLES, "match": "smart", "match_strictness": "normal",
        },
    }


def _build_category_filter(industries: list[str],
                            with_country: bool | list[str] = False) -> dict:
    """Empirically-verified shape (scripts/verify_prospeo_shape.py, 2026-05-14):
      - Industry key is `company_industry` (SINGULAR, top-level)
      - Values are LinkedIn-2023-style strings drawn verbatim from Prospeo's
        enum at https://prospeo.io/api-docs/enum/industries
      - Wrong shapes (filter ignored or rejected) include:
          * filters.company.industries.include — accepted but silently ignored
          * filters.company.industry.include — accepted but silently ignored
          * filters.industries.include — INVALID_FILTERS
          * filters.industry.include — INVALID_FILTERS
          * filters.company_industries.include (plural) — INVALID_FILTERS

    company_location_search behavior is undocumented for raw strings — kept
    behind --with-country flag pending a separate verification.
    """
    f: dict = {
        "company_industry": {"include": industries},
        "person_job_title": {
            "include": PILOT_TITLES, "match": "smart", "match_strictness": "normal",
        },
    }
    if with_country:
        countries = with_country if isinstance(with_country, list) else ["United States"]
        f["company_location_search"] = {"include": countries}
    return f


def load_sample_domains(n: int) -> list[str]:
    """Pull N random domains from inclusion_clean.csv (Website column).
    Reservoir sample so we don't hold the whole 27MB file in memory."""
    if not INCLUSION_CSV.exists():
        sys.exit(f"Missing {INCLUSION_CSV} — run scripts/clean_inclusion.py first.")
    sample: list[str] = []
    seen: set[str] = set()
    with INCLUSION_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        rdr = csv.DictReader(f)
        cols = {c.lower().strip(): c for c in (rdr.fieldnames or [])}
        key = cols.get("website") or cols.get("domain") or cols.get("url")
        if not key:
            sys.exit("inclusion_clean.csv missing Website/Domain/URL column")
        i = 0
        for row in rdr:
            d = _norm_domain(row.get(key))
            if not d or d in seen:
                continue
            seen.add(d)
            i += 1
            if len(sample) < n:
                sample.append(d)
            else:
                j = random.randint(0, i - 1)
                if j < n:
                    sample[j] = d
    return sample


def _process_results(results: list[dict], rows: list[dict],
                     api_key: str, llm_client: anthropic.Anthropic | None,
                     llm_prompt: str | None,
                     credits_already_spent: int, credits: int,
                     max_credits: int, target_results: int) -> int:
    """Enrich + filter every result; mutates `rows`. Returns updated credits."""
    for r in results:
        if len(rows) >= target_results:
            break
        if credits_already_spent + credits >= max_credits:
            break
        p = r.get("person") or {}
        c_block = r.get("company") or {}
        pid = p.get("person_id")
        if not pid:
            continue
        search_row = {
            "first_name": p.get("first_name"),
            "last_name": p.get("last_name"),
            "title": p.get("current_job_title"),
            "company_name": c_block.get("name") or c_block.get("company_name"),
            "company_website": c_block.get("website") or c_block.get("url"),
            "company_description": c_block.get("description"),
        }
        enriched, ec = _enrich_person(pid, api_key)
        credits += ec
        email = (enriched.get("email") or "").lower().strip()
        row = {
            "email": email or "",
            "first_name": enriched.get("first_name") or search_row["first_name"],
            "last_name": enriched.get("last_name") or search_row["last_name"],
            "title": enriched.get("title") or search_row["title"],
            "company_name": enriched.get("company_name") or search_row["company_name"],
            "company_website": enriched.get("company_website") or search_row["company_website"],
            "company_description": enriched.get("company_description") or search_row["company_description"],
            "enriched_with_email": bool(email),
        }
        rule = rule_classify(row)
        if rule:
            result, method, reason = rule
        elif email and llm_client and llm_prompt:
            result, reason = llm_classify_one(llm_client, llm_prompt, row)
            method = "llm"
        elif not email:
            result, method, reason = ("skipped", "no-email", "no verified email returned")
        else:
            result, method, reason = ("unknown", "none", "rule indeterminate + no llm")
        row.update(filter_result=result, filter_method=method, filter_reason=reason,
                   accepted=(result == "brand"))
        rows.append(row)
    return credits


def run_domain_pass(domains: list[str], api_key: str,
                    llm_client: anthropic.Anthropic | None, llm_prompt: str | None,
                    max_credits: int, credits_already_spent: int,
                    target_results: int) -> tuple[list[dict], int]:
    """Domain mode — batch domains in groups of 15 (Prospeo limit).
    Aborts a batch if Prospeo silently ignored the websites filter (returns
    more results than the cap)."""
    print(f"\n--- Pass: DOMAIN ({len(domains)} domains in batches of {PROSPEO_BATCH_DOMAINS}) ---")
    credits = 0
    rows: list[dict] = []
    batches = [domains[i:i + PROSPEO_BATCH_DOMAINS]
               for i in range(0, len(domains), PROSPEO_BATCH_DOMAINS)]

    for bi, batch in enumerate(batches, 1):
        if len(rows) >= target_results:
            break
        if credits_already_spent + credits >= max_credits:
            print(f"  ! budget cap hit ({credits_already_spent + credits}/"
                  f"{max_credits}) — stopping")
            break
        filters = {
            "company": {"websites": {"include": batch}},
            "person_job_title": {
                "include": PILOT_TITLES, "match": "smart", "match_strictness": "normal",
            },
        }
        data, c = _search_person(filters, api_key, page=1)
        credits += c
        pagination = data.get("pagination") or {}
        total_count = pagination.get("total_count", 0)
        results = data.get("results") or []
        # Safety: if Prospeo silently ignored the websites filter, this returns
        # a global result count way larger than possible for our small batch.
        if total_count > PROSPEO_MAX_RESULTS_PER_BATCH:
            print(f"  ! batch {bi}/{len(batches)} total_count={total_count} "
                  f"(websites filter likely ignored); skipping batch")
            continue
        print(f"  batch {bi}/{len(batches)}: {len(results)} results "
              f"(total_count={total_count}, credits={credits})")
        credits = _process_results(results, rows, api_key, llm_client, llm_prompt,
                                    credits_already_spent, credits, max_credits,
                                    target_results)
    return rows, credits


def run_category_pass(industries: list[str], api_key: str,
                      llm_client: anthropic.Anthropic | None, llm_prompt: str | None,
                      max_credits: int, credits_already_spent: int,
                      target_results: int, with_country: bool | list[str] = False
                      ) -> tuple[list[dict], int]:
    """Category mode — single filter, paginate until target_results hit or budget cap."""
    filters = _build_category_filter(industries, with_country=with_country)
    print(f"\n--- Pass: CATEGORY ({len(industries)} industries, "
          f"with_country={with_country}) ---")
    credits = 0
    rows: list[dict] = []
    page = 1
    while len(rows) < target_results:
        if credits_already_spent + credits >= max_credits:
            print(f"  ! budget cap hit — stopping")
            break
        data, c = _search_person(filters, api_key, page=page)
        credits += c
        pagination = data.get("pagination") or {}
        total_count = pagination.get("total_count", 0)
        results = data.get("results") or []
        print(f"  search-person page {page}: {len(results)} results "
              f"(total_count={total_count}, credits={credits})")
        credits = _process_results(results, rows, api_key, llm_client, llm_prompt,
                                    credits_already_spent, credits, max_credits,
                                    target_results)
        if page >= (pagination.get("total_page") or 1):
            break
        page += 1
    return rows, credits


# ---------- Output ----------

XLSX_COLS = [
    "mode", "filter_result", "filter_method", "accepted", "email",
    "first_name", "last_name", "title", "company_name", "company_website",
    "company_description", "enriched_with_email", "filter_reason",
]


def write_xlsx(domain_rows: list[dict], category_rows: list[dict],
               summary: dict, out_path: Path) -> None:
    wb = Workbook()
    bold_white = Font(bold=True, color="FFFFFF")
    blue_fill = PatternFill("solid", fgColor="0550AE")

    def fill(ws, title, rows):
        ws.title = title
        ws.append(XLSX_COLS)
        for cell in ws[1]:
            cell.font = bold_white
            cell.fill = blue_fill
        for r in rows:
            ws.append([
                "" if r.get(c) is None else str(r.get(c))[:300]
                for c in XLSX_COLS
            ])
        for col_idx, col_name in enumerate(XLSX_COLS, start=1):
            max_len = max((len(str(r.get(col_name, ""))) for r in rows), default=0)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
                min(max(len(col_name), max_len) + 2, 60)

    combined = (
        [{**r, "mode": "domain"} for r in domain_rows]
        + [{**r, "mode": "category"} for r in category_rows]
    )
    accepted_rows = [r for r in combined if r.get("accepted")]
    rejected_rows = [r for r in combined if not r.get("accepted")]

    ws_acc = wb.active
    fill(ws_acc, "Accepted", accepted_rows)
    ws_rej = wb.create_sheet("Rejected")
    fill(ws_rej, "Rejected", rejected_rows)

    ws_s = wb.create_sheet("Summary")
    ws_s.append(["metric", "value"])
    for k, v in summary.items():
        ws_s.append([k, v])

    wb.save(out_path)


def summarize(label: str, rows: list[dict]) -> dict:
    n = len(rows)
    accepted = sum(1 for r in rows if r.get("accepted"))
    rejected_by_llm = sum(1 for r in rows if r.get("filter_method") == "llm"
                          and not r.get("accepted"))
    rejected_by_rule = sum(1 for r in rows if r.get("filter_method") == "rule")
    no_email = sum(1 for r in rows if not r.get("enriched_with_email"))
    return {
        f"{label} - total_results": n,
        f"{label} - accepted (brand)": accepted,
        f"{label} - brand_rate_%": round(100 * accepted / n, 1) if n else 0.0,
        f"{label} - rejected_by_rule": rejected_by_rule,
        f"{label} - rejected_by_llm": rejected_by_llm,
        f"{label} - no_email_returned": no_email,
    }


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--max-credits", type=int, default=DEFAULT_MAX_CREDITS,
                    help=f"Hard budget cap (default {DEFAULT_MAX_CREDITS} ~= ${DEFAULT_MAX_CREDITS * 0.02:.2f})")
    ap.add_argument("--dry-run", action="store_true",
                    help="Show plan + sample inputs; make no API calls")
    ap.add_argument("--target-results", type=int, default=DOMAIN_SAMPLE_SIZE,
                    help="Target accepted+rejected leads per pass")
    ap.add_argument("--skip-llm", action="store_true",
                    help="Don't instantiate Anthropic client (rule-only filter)")
    ap.add_argument("--yes", action="store_true",
                    help="Skip the confirmation prompt before live API calls")
    ap.add_argument("--skip-domain", action="store_true",
                    help="Skip Pass A (domain mode) entirely. Runs only Pass B "
                         "(category mode). Useful once category mode is the "
                         "primary path and you only want a category sample.")
    ap.add_argument("--with-country", nargs="?", const="United States", default=None,
                    metavar="LIST",
                    help="Add company_location_search to category filter. "
                         "Pass alone for United States only, or with a comma-separated "
                         "list (e.g. --with-country 'United States,Canada'). "
                         "Off by default — Prospeo may reject raw country strings.")
    args = ap.parse_args()

    # Normalize --with-country to a list (or False if not passed)
    if args.with_country is None:
        args.with_country = False
    else:
        args.with_country = [c.strip() for c in args.with_country.split(",") if c.strip()]

    load_dotenv()
    api_key = os.environ.get("PROSPEO_API_KEY", "").strip()
    if not api_key and not args.dry_run:
        sys.exit("PROSPEO_API_KEY not set in .env")

    sample_domains: list[str] = [] if args.skip_domain else load_sample_domains(DOMAIN_SAMPLE_SIZE)

    print("=" * 70)
    print("PROSPEO CATEGORY-vs-DOMAIN PILOT")
    print("=" * 70)
    if args.skip_domain:
        print("  Domain pass:          SKIPPED (--skip-domain)")
    else:
        print(f"  Domain sample size:   {len(sample_domains)} (from inclusion_clean.csv)")
    print(f"  Category list:        {len(PILOT_INDUSTRIES)} industries")
    for ind in PILOT_INDUSTRIES:
        print(f"      - {ind}")
    print(f"  Title filter:         {len(PILOT_TITLES)} titles (owner + marketing/e-com)")
    print(f"  Target results / pass: {args.target_results}")
    print(f"  Max credits:          {args.max_credits} (~=${args.max_credits * 0.02:.2f})")
    print()
    if not args.skip_domain:
        print("  First 5 sample domains for Pass A:")
        for d in sample_domains[:5]:
            print(f"    - {d}")
        print()

    if args.dry_run:
        print("=== DRY RUN -- no API calls made ===")
        print()
        if args.skip_domain:
            print("Pass A (Domain): SKIPPED")
        else:
            n_batches = (len(sample_domains) + PROSPEO_BATCH_DOMAINS - 1) // PROSPEO_BATCH_DOMAINS
            print(f"Pass A (Domain): {len(sample_domains)} domains -> "
                  f"{n_batches} batch(es) of {PROSPEO_BATCH_DOMAINS} max")
            print("First batch filter:")
            ex_domains = sample_domains[:PROSPEO_BATCH_DOMAINS]
            ex_filter = {
                "company": {"websites": {"include": ex_domains}},
                "person_job_title": {
                    "include": PILOT_TITLES, "match": "smart", "match_strictness": "normal",
                },
            }
            print(f"  {json.dumps(ex_filter, indent=2)}")
            print()
        print("Pass B (Category) filter:")
        print(f"  {json.dumps(_build_category_filter(PILOT_INDUSTRIES, with_country=args.with_country), indent=2)}")
        return

    if not args.yes:
        ans = input(f"\nReady to spend up to {args.max_credits} credits "
                    f"(~${args.max_credits * 0.02:.2f}). Continue? [y/N] ").strip().lower()
        if ans != "y":
            sys.exit("Aborted by user.")

    llm_client = None
    llm_prompt = None
    if not args.skip_llm:
        llm_client = anthropic.Anthropic()
        llm_prompt = AGENCY_FILTER_PROMPT.read_text(encoding="utf-8")

    spent = 0
    if args.skip_domain:
        domain_rows: list[dict] = []
        c1 = 0
        print("  Pass A (Domain) skipped via --skip-domain.")
    else:
        domain_rows, c1 = run_domain_pass(
            sample_domains, api_key, llm_client, llm_prompt,
            args.max_credits, spent, args.target_results,
        )
        spent += c1
        print(f"  Pass A used {c1} credits; total now {spent}/{args.max_credits}")

    category_rows: list[dict] = []
    c2 = 0
    if spent < args.max_credits:
        category_rows, c2 = run_category_pass(
            PILOT_INDUSTRIES, api_key, llm_client, llm_prompt,
            args.max_credits, spent, args.target_results,
            with_country=args.with_country,
        )
        spent += c2
        print(f"  Pass B used {c2} credits; total now {spent}/{args.max_credits}")
    else:
        print("  Skipping Pass B — budget already exhausted by Pass A.")

    summary: dict = {}
    summary.update(summarize("Pass A (Domain)", domain_rows))
    summary.update(summarize("Pass B (Category)", category_rows))
    summary["Total credits spent"] = spent
    summary["Total cost USD"] = round(spent * 0.02, 2)

    print("\n=== Pilot summary ===")
    for k, v in summary.items():
        print(f"  {k}: {v}")

    os.makedirs("exports", exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    out = Path("exports") / f"category_pilot_{stamp}.xlsx"
    write_xlsx(domain_rows, category_rows, summary, out)
    print(f"\nWrote {out}")


if __name__ == "__main__":
    main()
