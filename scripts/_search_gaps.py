"""Search the xlsx for industries matching the set-1 gaps."""

import re
from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent / "exports" / "prospeo_industry_candidates.xlsx"

GAPS = {
    "supplements": re.compile(r"supplement|vitamin|nutraceutical", re.I),
    "alt_medicine": re.compile(r"alternative medicine|herbal|naturopath|holistic|homeopath|ayurved", re.I),
    "groceries": re.compile(r"grocer|specialty food|natural food", re.I),
    "health_wellness": re.compile(r"wellness|holistic health", re.I),
    "mechanical_goods": re.compile(r"mechanical|machinery|hardware|tools", re.I),
    "electronics": re.compile(r"\belectronic|consumer electronic|gadget", re.I),
}

wb = load_workbook(XLSX, read_only=True)

for sheet_name in ("all_industries", "positive_industries", "ecom_industries"):
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    i_ind = headers.index("industry")
    i_n = headers.index("lead_count")

    print(f"\n=== {sheet_name} ({len(rows)-1} rows) ===")
    for gap, pat in GAPS.items():
        hits = []
        for r in rows[1:]:
            if not r or not r[i_ind]:
                continue
            if pat.search(str(r[i_ind])):
                hits.append((r[i_n] or 0, str(r[i_ind])))
        hits.sort(key=lambda x: -x[0])
        print(f"\n  [{gap}] {len(hits)} matches")
        for n, name in hits[:8]:
            print(f"    {n:>6}  {name}")
        if len(hits) > 8:
            print(f"    ... +{len(hits) - 8} more")
