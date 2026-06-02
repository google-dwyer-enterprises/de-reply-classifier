"""READ-ONLY: distinct Apollo industries in our lead pool, with a Prospeo-compat hint.

Writes exports/prospeo_industry_candidates.xlsx with two sheets:
  - all_industries      : every distinct lead_contacts.industry + lead_count
  - positive_industries : same but restricted to leads whose status (coalesce of
                          manual_status, auto_status) is booked / interested /
                          interested_past, with per-status counts.

Both sheets carry:
  - prospeo_compatible:    accepted | unknown | empty
  - suggested_prospeo_value: conservative manual renames only; blank otherwise.

Pair with `scripts/verify_prospeo_shape.py --from-xlsx <this file>` to probe
unknown values against Prospeo's live API (rejections are free).
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook

from db import connect


VERIFIED_ACCEPTED = {
    "Retail Apparel and Fashion",
    "Apparel Manufacturing",
    "Cosmetics",
    "Personal Care Product Manufacturing",
    "Food and Beverage Manufacturing",
    "Furniture and Home Furnishings Manufacturing",
    "Sporting Goods Manufacturing",
    "Consumer Goods",
    "Pet Services",
}

# Conservative manual renames: Apollo (pre-2023 LinkedIn) -> Prospeo verified-9.
# Only obvious, unambiguous mappings. Everything else stays blank for human review.
MANUAL_RENAMES = {
    "Apparel & Fashion": "Retail Apparel and Fashion",
    "Apparel and Fashion": "Retail Apparel and Fashion",
    "Sporting Goods": "Sporting Goods Manufacturing",
    "Consumer Goods": "Consumer Goods",
    "Cosmetics": "Cosmetics",
}

POSITIVE_STATUSES = ("booked", "interested", "interested_past")

ALL_SQL = """
select
  industry,
  count(*)::int as lead_count
from lead_contacts
where industry is not null and btrim(industry) <> ''
group by industry
order by count(*) desc, industry asc
"""

POSITIVE_SQL = """
with eff as (
  select
    l.lead_email,
    coalesce(l.manual_status, l.auto_status) as status
  from leads l
  where coalesce(l.manual_status, l.auto_status) in %s
)
select
  lc.industry,
  count(*)::int as lead_count,
  count(*) filter (where eff.status = 'booked')::int          as booked,
  count(*) filter (where eff.status = 'interested')::int      as interested,
  count(*) filter (where eff.status = 'interested_past')::int as interested_past
from eff
join lead_contacts lc on lc.lead_email = eff.lead_email
where lc.industry is not null and btrim(lc.industry) <> ''
group by lc.industry
order by count(*) desc, lc.industry asc
"""


def compat_for(industry: str | None) -> str:
    if not industry or not industry.strip():
        return "empty"
    if industry in VERIFIED_ACCEPTED:
        return "accepted"
    return "unknown"


def suggested_for(industry: str | None) -> str:
    if not industry:
        return ""
    if industry in VERIFIED_ACCEPTED:
        return industry
    return MANUAL_RENAMES.get(industry, "")


def write_sheet(ws, header: list[str], rows: list[tuple]) -> None:
    ws.append(header)
    for row in rows:
        ws.append(list(row))
    # Auto-ish width
    for col_idx, name in enumerate(header, start=1):
        max_len = max(
            [len(str(name))] + [len(str(r[col_idx - 1])) for r in rows] or [0]
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_len + 2, 60
        )


def main() -> None:
    conn = connect()
    try:
        with conn.cursor() as cur:
            cur.execute(ALL_SQL)
            all_rows = cur.fetchall()
            cur.execute(POSITIVE_SQL, (POSITIVE_STATUSES,))
            pos_rows = cur.fetchall()
    finally:
        conn.close()

    print(f"all_industries:       {len(all_rows)} distinct")
    print(f"positive_industries:  {len(pos_rows)} distinct "
          f"(status in {POSITIVE_STATUSES})")

    all_out = [
        (ind, n, compat_for(ind), suggested_for(ind))
        for (ind, n) in all_rows
    ]
    pos_out = [
        (ind, n, b, i, ip, compat_for(ind), suggested_for(ind))
        for (ind, n, b, i, ip) in pos_rows
    ]

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "all_industries"
    write_sheet(
        ws_all,
        ["industry", "lead_count", "prospeo_compatible", "suggested_prospeo_value"],
        all_out,
    )

    ws_pos = wb.create_sheet("positive_industries")
    write_sheet(
        ws_pos,
        ["industry", "lead_count", "booked", "interested", "interested_past",
         "prospeo_compatible", "suggested_prospeo_value"],
        pos_out,
    )

    out_dir = Path(__file__).resolve().parent.parent / "exports"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "prospeo_industry_candidates.xlsx"
    wb.save(out_path)
    print(f"\nWrote {out_path}")

    # Quick console summary of compat split
    def split(rows):
        c = {"accepted": 0, "unknown": 0, "empty": 0}
        for r in rows:
            c[compat_for(r[0])] += 1
        return c

    print(f"  all      compat split: {split(all_rows)}")
    print(f"  positive compat split: {split(pos_rows)}")


if __name__ == "__main__":
    main()
