"""One-shot: write probe_results sheet from a manually-supplied accepted list.
The remaining 'unknown' industries are marked rejected (matches the user's
"4 accepted, all others rejected" report)."""

from pathlib import Path
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent / "exports" / "prospeo_industry_candidates.xlsx"

ACCEPTED = {
    "Automotive",
    "Computer Hardware",
    "Consumer Services",
    "Human Resources",
}

wb = load_workbook(XLSX)

unknowns = set()
ws = wb["positive_industries"]
rows = list(ws.iter_rows(values_only=True))
headers = [str(h) if h is not None else "" for h in rows[0]]
i_ind = headers.index("industry")
i_compat = headers.index("prospeo_compatible")
for r in rows[1:]:
    if r and r[i_compat] == "unknown" and r[i_ind]:
        unknowns.add(str(r[i_ind]))

results = []
for v in sorted(unknowns):
    if v in ACCEPTED:
        results.append((v, "accepted", 200, None, ""))
    else:
        results.append((v, "rejected", 400, None,
                        f"Invalid industry(s) in include list: ['{v}']"))

if "probe_results" in wb.sheetnames:
    del wb["probe_results"]
ws = wb.create_sheet("probe_results")
ws.append(["industry", "verdict", "http_status", "total_count", "message"])
for r in results:
    ws.append(list(r))

wb.save(XLSX)
print(f"Wrote {len(results)} rows to probe_results in {XLSX}")
print(f"  accepted: {sum(1 for r in results if r[1] == 'accepted')}")
print(f"  rejected: {sum(1 for r in results if r[1] == 'rejected')}")
