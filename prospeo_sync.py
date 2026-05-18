"""Prospeo lead scraper.

Pulls decision-maker leads from Prospeo for each domain in domain_inclusion_list,
dedupes against existing lead_contacts/replies/prospeo_new_leads, applies a
rule-based brand/agency/reseller filter (with an LLM grey-zone pass), and writes
accepted leads to prospeo_new_leads + a CSV export for Jam.

CLI: python run.py scrape-leads [--domains <csv>] [--limit N] [--dry-run] [--skip-llm]

Prereqs:
  - PROSPEO_API_KEY in .env
  - domain_inclusion_list populated (or pass --domains <csv>)

This is a v1 scaffold. The Prospeo HTTP call is stubbed (`_prospeo_search`) until
we confirm the exact API shape from their docs + credentials.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import anthropic
import psycopg2
import requests
from dotenv import load_dotenv
from psycopg2.extras import execute_values

from db import connect


# --- config -----------------------------------------------------------------

PROSPEO_BASE = "https://api.prospeo.io"
PROSPEO_TIMEOUT = 30
# Prospeo silently ignores the websites filter past ~15 entries (undocumented).
# Their docs claim 500 max, but empirically the filter falls back to "no
# websites filter" above 15, returning unfiltered global results. Verified by
# manual testing 2026-05-13 after a 506-credit burn. Keep this at 15.
PROSPEO_BATCH_DOMAINS = 15
# With 15 domains per batch and owner-only titles, legit batches return at most
# ~5 results = 1 page. Pagination is only useful if the filter misbehaves, so
# we cap aggressively.
PROSPEO_MAX_PAGES_PER_BATCH = 3
# Safety: if a batch returns more results than possible for its size, the
# website filter has been silently ignored — abort that batch.
PROSPEO_MAX_RESULTS_PER_BATCH = PROSPEO_BATCH_DOMAINS * 5

# Decision-maker job titles (Prospeo's Smart-match resolves variants).
#
# Owner-only list derived from Dwyer reply data (see prospeo_sync analysis
# in conversation 2026-05-13): marketing/CMO/e-com-manager titles produced
# 1 booking out of 15 replies in v3 classifications, while
# Founder/CEO/Owner/President generated all real positive outcomes.
# At small/mid DTC brands the owner IS the buyer for agency services.
PROSPEO_TITLES = [
    # Tier 1: owner (15.8% lead-level conversion per title_analysis 2026-05-15)
    "CEO",
    "Chief Executive Officer",
    "Founder",
    "Co-Founder",
    "Owner",
    "Co-Owner",
    "President",
    "Founder and CEO",
    "CEO and Founder",
    "Managing Director",
    "Chairman",
    # Tier 2: marketing/e-com (8.6% lead-level conversion, n=315)
    "CMO",
    "Chief Marketing Officer",
    "Head of Marketing",
    "VP Marketing",
    "Head of E-commerce",
    "Director of E-commerce",
]

# Reference: previous broader list (kept for re-tuning runs).
# Marketing/e-com-specialist titles dropped after reply-data analysis.
_PROSPEO_TITLES_BROAD_REF = [
    "CEO", "Founder", "Co-Founder", "Owner", "President",
    "CMO", "VP Marketing", "Head of Marketing", "Marketing Director",
    "Head of E-commerce", "Chief Marketing Officer", "Chief E-commerce Officer",
    "VP of E-commerce", "Director of E-commerce",
]

# E-commerce-relevant industry strings — EMPIRICALLY VERIFIED against
# Prospeo's live API on 2026-05-14 + 2026-05-15 via
# scripts/verify_prospeo_shape.py. Each was confirmed to (1) parse without
# INVALID_FILTERS and (2) reduce the global total_count when applied
# alongside the title filter.
#
# Prospeo uses LinkedIn's POST-2023 taxonomy. Common pitfalls that DO NOT
# work (verified rejected): "Apparel and Fashion", "Retail", "Furniture",
# "Manufacturing", "Apparel & Fashion" (ampersand), "Personal Care Products",
# "Health and Beauty", "Food Production", "Beverage Manufacturing".
#
# Used by scripts/prospeo_category_pilot.py and (Step 4 onward) by
# prospeo_sync.run() under --mode category.
PROSPEO_INDUSTRIES = [
    # Verified 2026-05-14 (initial 9)
    "Retail Apparel and Fashion",
    "Apparel Manufacturing",
    "Cosmetics",
    "Personal Care Product Manufacturing",
    "Food and Beverage Manufacturing",
    "Furniture and Home Furnishings Manufacturing",
    "Sporting Goods Manufacturing",
    "Consumer Goods",
    "Pet Services",
    # Verified 2026-05-15 (3 additions from ecom + gaps probes)
    "Retail Groceries",                          # probe on ecom_industries sheet
    "Alternative Medicine",                      # gaps probe, total_count=128K
    "Retail Health and Personal Care Products",  # gaps probe, total_count=141K
]

DECISION_MAKER_TITLES = {
    # exact-match (lowercased) keywords; we check substring containment
    "ceo", "chief executive",
    "founder", "co-founder", "cofounder",
    "owner", "president",
    "cmo", "chief marketing",
    "vp marketing", "vp of marketing", "head of marketing", "director of marketing",
    "head of e-commerce", "head of ecommerce",
    "chief e-commerce", "chief ecommerce",
    "vp e-commerce", "vp ecommerce", "vp of ecommerce",
    "director of e-commerce", "director of ecommerce",
}

# Hard rule: if company name OR website contains any of these tokens → agency
AGENCY_TOKENS = {
    "agency", "marketing", "consulting", "consultancy", "services",
    "solutions", "media", "digital", "growth", "fulfillment", "3pl",
    "partners", "advisors", "advisory",
}

# Known marketplaces — auto-reject
MARKETPLACE_DOMAINS = {
    "amazon.com", "walmart.com", "etsy.com", "ebay.com",
    "faire.com", "shopify.com", "alibaba.com", "aliexpress.com",
}

AGENCY_FILTER_MODEL = "claude-haiku-4-5"
AGENCY_FILTER_PROMPT = Path(__file__).parent / "prompts" / "agency_filter.txt"


# --- helpers ----------------------------------------------------------------

# Common multi-part TLDs we should preserve (e.g. brand.com.au, brand.co.uk).
# Without this list, `awesomebrand.com.uy` would be stripped to `com.uy` and
# rejected by Prospeo as "Invalid website format".
_MULTI_PART_TLDS = {
    # .co.*
    "co.uk", "co.nz", "co.jp", "co.kr", "co.za", "co.in", "co.il", "co.id",
    "co.th", "co.ke", "co.ug", "co.tz", "co.ma", "co.ve", "co.cr",
    # .com.*
    "com.au", "com.br", "com.mx", "com.sg", "com.hk", "com.tw", "com.tr",
    "com.ar", "com.co", "com.pe", "com.ph", "com.my", "com.cn", "com.uy",
    "com.ng", "com.pk", "com.ec", "com.ve", "com.do", "com.bo", "com.gt",
    "com.pa", "com.sv", "com.ni", "com.eg", "com.sa", "com.lb", "com.ua",
    # .org.*
    "org.uk", "org.il", "org.sg", "org.au", "org.nz", "org.za",
    # .net.*
    "net.au", "net.nz", "net.br",
    # .gov.* / .ac.* / .edu.*
    "gov.uk", "gov.au", "ac.uk", "ac.nz", "edu.au", "edu.sg",
    # Canadian provinces (geographic subdomains under .ca)
    "bc.ca", "ab.ca", "on.ca", "qc.ca", "ns.ca", "nb.ca", "mb.ca", "sk.ca",
    "pe.ca", "nl.ca", "yk.ca", "nt.ca", "nu.ca",
    # Australian states
    "nsw.au", "vic.au", "qld.au", "sa.au", "wa.au", "tas.au", "act.au",
    "nt.au",
}


def _norm_domain(d: str | None) -> str | None:
    """Normalize a URL/domain to a registrable domain. Strips scheme, www, path,
    AND subdomains. parts.agcocorp.com → agcocorp.com.
    Preserves multi-part TLDs like .com.au, .co.uk."""
    if not d:
        return None
    d = d.strip().lower()
    if d.startswith("http://"):
        d = d[7:]
    if d.startswith("https://"):
        d = d[8:]
    if d.startswith("www."):
        d = d[4:]
    # strip path, query, fragment
    for sep in ("/", "?", "#"):
        if sep in d:
            d = d.split(sep, 1)[0]
    if not d:
        return None
    # Strip subdomains (Prospeo rejects them with INVALID_FILTERS)
    parts = d.split(".")
    if len(parts) > 2:
        last_two = ".".join(parts[-2:])
        if last_two in _MULTI_PART_TLDS and len(parts) >= 3:
            # brand.co.uk → keep last 3
            d = ".".join(parts[-3:])
        else:
            # parts.agcocorp.com → agcocorp.com
            d = ".".join(parts[-2:])
    return d


def _email_domain(email: str | None) -> str | None:
    if not email or "@" not in email:
        return None
    return email.rsplit("@", 1)[1].lower()


# --- domain sourcing --------------------------------------------------------

DOMAIN_COL_CANDIDATES = ("domain", "website", "company website", "url", "site")


def load_domains_from_csv(path: str) -> list[str]:
    """Accepts CSV or tab-separated. Prefers a column named domain/website/url/site;
    falls back to the first column. Deduplicates while preserving first-seen order."""
    # Sniff delimiter (Jam's exports are tab-separated; standard CSV also supported)
    with open(path, newline="", encoding="utf-8") as f:
        sample = f.read(4096)
        f.seek(0)
        delim = "\t" if sample.count("\t") > sample.count(",") else ","
        rdr = csv.DictReader(f, delimiter=delim)
        key = None
        if rdr.fieldnames:
            lower_map = {h.lower().strip(): h for h in rdr.fieldnames}
            for cand in DOMAIN_COL_CANDIDATES:
                if cand in lower_map:
                    key = lower_map[cand]
                    break
        out: list[str] = []
        seen: set[str] = set()
        if key:
            for row in rdr:
                d = _norm_domain(row.get(key))
                if d and d not in seen:
                    seen.add(d)
                    out.append(d)
        else:
            f.seek(0)
            for line in f:
                d = _norm_domain(line.strip().split(delim)[0])
                if d and d not in seen:
                    seen.add(d)
                    out.append(d)
    return out


def load_domains_from_db(conn, limit: int | None = None) -> list[str]:
    with conn.cursor() as cur:
        sql = (
            "select domain from domain_inclusion_list "
            "order by coalesce(last_scraped_at, 'epoch'::timestamptz) asc, domain"
        )
        if limit:
            sql += f" limit {int(limit)}"
        cur.execute(sql)
        return [r[0] for r in cur.fetchall()]


def fetch_existing_emails(conn) -> set[str]:
    """Emails we already have anywhere — used to dedupe Prospeo results."""
    out: set[str] = set()
    with conn.cursor() as cur:
        cur.execute("select lead_email from lead_contacts where lead_email is not null")
        out.update(r[0].lower() for r in cur.fetchall() if r[0])
        cur.execute("select distinct lead_email from replies")
        out.update(r[0].lower() for r in cur.fetchall() if r[0])
        cur.execute("select email from prospeo_new_leads")
        out.update(r[0].lower() for r in cur.fetchall() if r[0])
    return out


def fetch_domains_with_decision_maker(conn) -> set[str]:
    """Domains where we already have at least one decision-maker contact —
    skip these so we don't re-pay for emails we already have.

    The 230k-row scan occasionally drops the Supabase pooler connection
    ('SSL connection has been closed unexpectedly'). On failure we open a
    fresh short-lived connection just for this read and retry up to 3 times.
    The caller's conn is left alone so subsequent queries can use it.
    """
    sql = """
      select distinct lower(split_part(lead_email, '@', 2)) as d
      from lead_contacts
      where lead_email is not null
        and lower(coalesce(title, '')) ~ '(ceo|founder|owner|president|chief|cmo|head of (marketing|e-?commerce))'
    """
    import time
    last_exc: Exception | None = None
    use_conn = conn
    owns_conn = False
    for attempt in range(3):
        try:
            with use_conn.cursor() as cur:
                cur.execute(sql)
                rows = {r[0] for r in cur.fetchall() if r[0]}
            if owns_conn:
                use_conn.close()
            return rows
        except psycopg2.OperationalError as exc:
            last_exc = exc
            print(f"  ! fetch_domains_with_decision_maker dropped "
                  f"({type(exc).__name__}); retrying with fresh connection "
                  f"({attempt + 1}/3)", file=sys.stderr)
            if owns_conn:
                try:
                    use_conn.close()
                except Exception:
                    pass
            time.sleep(2 * (attempt + 1))
            from db import connect as _reconnect
            use_conn = _reconnect()
            owns_conn = True
    raise last_exc  # type: ignore[misc]


def fetch_recently_scraped_domains(conn, stale_days: int = 30) -> set[str]:
    """Domains we've already scraped recently — skip on this run."""
    sql = f"""
      select domain from domain_inclusion_list
      where last_scraped_at is not null
        and last_scraped_at > now() - interval '{int(stale_days)} days'
    """
    with conn.cursor() as cur:
        cur.execute(sql)
        return {r[0] for r in cur.fetchall() if r[0]}


def mark_domains_scraped(conn, domains: list[str]) -> None:
    """Upsert a row in domain_inclusion_list for each domain, setting
    last_scraped_at = now(). Idempotent."""
    if not domains:
        return
    sql = """
      insert into domain_inclusion_list (domain, last_scraped_at)
      values (%s, now())
      on conflict (domain) do update set last_scraped_at = excluded.last_scraped_at
    """
    with conn.cursor() as cur:
        cur.executemany(sql, [(d,) for d in domains])
    conn.commit()


def fetch_category_state(conn) -> dict[str, dict]:
    """Read category_scrape_state. Returns {industry: state_dict}.

    Industries with no row yet are simply absent — caller treats absence
    as 'last_page_consumed=0, exhausted=False'.
    """
    sql = """
      select industry, countries, last_page_consumed, total_pages,
             exhausted, last_scraped_at, total_credits_spent
      from category_scrape_state
    """
    out: dict[str, dict] = {}
    with conn.cursor() as cur:
        cur.execute(sql)
        for row in cur.fetchall():
            ind, countries, last_page, total_pages, exhausted, last_scraped, credits = row
            out[ind] = {
                "countries": list(countries or []),
                "last_page_consumed": last_page or 0,
                "total_pages": total_pages,
                "exhausted": bool(exhausted),
                "last_scraped_at": last_scraped,
                "total_credits_spent": credits or 0,
            }
    return out


def upsert_category_state(conn, industry: str, countries: list[str],
                           last_page_consumed: int, total_pages: int | None,
                           exhausted: bool, credits_added: int) -> None:
    """Insert or update the state row for one industry. Idempotent.

    credits_added is added to total_credits_spent (cumulative across runs).
    last_scraped_at is always set to now().
    """
    sql = """
      insert into category_scrape_state (
        industry, countries, last_page_consumed, total_pages,
        exhausted, last_scraped_at, total_credits_spent
      )
      values (%s, %s, %s, %s, %s, now(), %s)
      on conflict (industry) do update set
        countries = excluded.countries,
        last_page_consumed = excluded.last_page_consumed,
        total_pages = excluded.total_pages,
        exhausted = excluded.exhausted,
        last_scraped_at = excluded.last_scraped_at,
        total_credits_spent = category_scrape_state.total_credits_spent + %s
    """
    with conn.cursor() as cur:
        cur.execute(sql, (industry, countries, last_page_consumed, total_pages,
                          exhausted, credits_added, credits_added))
    conn.commit()


# --- Prospeo API (stub) -----------------------------------------------------

def _post(path: str, body: dict, api_key: str) -> dict:
    url = f"{PROSPEO_BASE}{path}"
    headers = {"X-KEY": api_key, "Content-Type": "application/json"}
    resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    if resp.status_code == 429:
        import time; time.sleep(5)
        resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    resp.raise_for_status()
    return resp.json() or {}


def _search_people(domains: list[str], api_key: str) -> tuple[list[dict], int]:
    """Search-Person across a batch of domains. Filters server-side to
    decision-maker job titles. Returns (rows, credits_used).

    Cost: 1 credit per page fetched (whether it returns 25 or fewer results).
    Prospeo returns 400 / NO_RESULTS when a valid query matches nothing — we
    treat that as an empty result set, not a failure.
    """
    filters = {
        "company": {"websites": {"include": domains}},
        "person_job_title": {
            "include": PROSPEO_TITLES,
            "match": "smart",
            "match_strictness": "normal",
        },
    }
    out: list[dict] = []
    credits_used = 0
    page = 1
    url = f"{PROSPEO_BASE}/search-person"
    headers = {"X-KEY": api_key, "Content-Type": "application/json"}
    while page <= PROSPEO_MAX_PAGES_PER_BATCH:
        resp = requests.post(url, json={"filters": filters, "page": page},
                             headers=headers, timeout=PROSPEO_TIMEOUT)
        if resp.status_code == 429:
            import time; time.sleep(5)
            resp = requests.post(url, json={"filters": filters, "page": page},
                                 headers=headers, timeout=PROSPEO_TIMEOUT)
        # Treat 400/NO_RESULTS as empty success (no credits charged)
        if resp.status_code == 400:
            try:
                payload = resp.json()
            except ValueError:
                payload = {}
            if payload.get("error_code") == "NO_RESULTS":
                break
            print(f"  ! search-person 400: {resp.text[:200]}", file=sys.stderr)
            break
        resp.raise_for_status()
        data = resp.json() or {}
        credits_used += 1  # Prospeo charges 1 credit per page fetched
        if data.get("error"):
            print(f"  ! search-person error: {data}", file=sys.stderr)
            break
        # Safety: detect Prospeo silently ignoring the websites filter.
        pagination = data.get("pagination") or {}
        total = pagination.get("total_count") or 0
        if total > PROSPEO_MAX_RESULTS_PER_BATCH:
            print(f"  ! websites filter ignored (total_count={total} for "
                  f"{len(domains)} domains); aborting batch to prevent credit burn",
                  file=sys.stderr)
            break
        for r in data.get("results") or []:
            p = r.get("person") or {}
            c = r.get("company") or {}
            out.append({
                "person_id": p.get("person_id"),
                "first_name": p.get("first_name"),
                "last_name": p.get("last_name"),
                "title": p.get("current_job_title"),
                "company_name": c.get("name") or c.get("company_name"),
                "company_website": c.get("website") or c.get("url"),
                "company_description": c.get("description"),
            })
        pagination = data.get("pagination") or {}
        if page >= (pagination.get("total_page") or 1):
            break
        page += 1
    return out, credits_used


class InsufficientCreditsError(Exception):
    """Prospeo returned INSUFFICIENT_CREDITS — top up your account.

    Caller should abort the run immediately. Retrying inside the same run
    will just produce more INSUFFICIENT_CREDITS responses (zero progress
    means the next call has the same problem). State is intentionally not
    touched on this error so that after top-up the run resumes cleanly.
    """
    pass


def _search_people_by_industry(industry: str, countries: list[str],
                                page: int, api_key: str
                                ) -> tuple[list[dict], int, int]:
    """Single-page Search-Person filtered by industry + titles + country.

    Returns (results, total_pages, credits_used).

    Cost: 1 credit per page fetched. Prospeo's 400/NO_RESULTS is treated as
    an empty result set with zero credits charged (matches _search_people).

    Unlike domain mode, category mode is paginated per industry across runs.
    The caller is responsible for tracking `last_page_consumed` in
    category_scrape_state and passing `page = last_page_consumed + 1`.
    """
    filters: dict = {
        "company_industry": {"include": [industry]},
        "person_job_title": {
            "include": PROSPEO_TITLES,
            "match": "smart",
            "match_strictness": "normal",
        },
    }
    if countries:
        filters["company_location_search"] = {"include": list(countries)}

    url = f"{PROSPEO_BASE}/search-person"
    headers = {"X-KEY": api_key, "Content-Type": "application/json"}
    body = {"filters": filters, "page": page}

    resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    if resp.status_code == 429:
        import time; time.sleep(5)
        resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)

    # Distinguish three distinct 400s:
    #   NO_RESULTS         -> valid empty result, page IS consumed (state advances)
    #   INSUFFICIENT_CREDITS -> hard abort, do NOT advance state (so resume works
    #                          cleanly after top-up)
    #   anything else      -> transient error, propagate to caller's retry path
    if resp.status_code == 400:
        try:
            payload = resp.json()
        except ValueError:
            payload = {}
        error_code = payload.get("error_code")
        if error_code == "NO_RESULTS":
            return ([], 0, 0)
        if error_code == "INSUFFICIENT_CREDITS":
            raise InsufficientCreditsError(
                f"Prospeo returned INSUFFICIENT_CREDITS (industry={industry!r}, "
                "page={page}). Top up your account at https://prospeo.io/dashboard "
                "and re-run — state is unchanged."
            )
        # Other 400s (INVALID_FILTERS, etc.) — raise so caller's
        # RequestException handler retries next cycle without touching state.
        print(f"  ! search-person 400 (industry={industry!r}): {resp.text[:200]}",
              file=sys.stderr)
        resp.raise_for_status()

    resp.raise_for_status()
    data = resp.json() or {}
    if data.get("error"):
        print(f"  ! search-person error (industry={industry!r}): {data}", file=sys.stderr)
        return ([], 0, 0)

    pagination = data.get("pagination") or {}
    total_pages = pagination.get("total_page") or 1
    out: list[dict] = []
    for r in data.get("results") or []:
        p = r.get("person") or {}
        c = r.get("company") or {}
        out.append({
            "person_id": p.get("person_id"),
            "first_name": p.get("first_name"),
            "last_name": p.get("last_name"),
            "title": p.get("current_job_title"),
            "company_name": c.get("name") or c.get("company_name"),
            "company_website": c.get("website") or c.get("url"),
            "company_description": c.get("description"),
        })
    return out, total_pages, 1


def _enrich_mobile(person_id: str, api_key: str) -> dict:
    """Enrich a previously-found person for mobile only (10 credits per verified
    mobile; free if not available). Returns {} on failure / no mobile."""
    url = f"{PROSPEO_BASE}/enrich-person"
    headers = {"X-KEY": api_key, "Content-Type": "application/json"}
    body = {
        "data": {"person_id": person_id},
        "enrich_mobile": True,
        "only_verified_mobile": True,
    }
    resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    if resp.status_code == 400:
        return {}  # no verified mobile available
    if resp.status_code == 429:
        import time; time.sleep(5)
        resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    resp.raise_for_status()
    data = resp.json() or {}
    if data.get("error"):
        return {}
    person = data.get("person") or {}
    mob = person.get("mobile") or {}
    mobile_value = mob.get("mobile") or mob.get("mobile_international")
    # 10 credits if a mobile was returned and Prospeo billed it
    billed = 0 if data.get("free_enrichment") else (10 if mobile_value else 0)
    return {
        "mobile": mobile_value,
        "mobile_status": mob.get("status"),
        "_credits": billed,
    }


def _enrich_person(person_id: str, api_key: str) -> dict:
    """Returns enriched person dict with email (or None if UNAVAILABLE).
    Cost: 1 credit if an email is returned; free otherwise.
    With only_verified_email=True, Prospeo returns 400/NO_MATCH for persons
    without a verified email — we treat that as a soft skip."""
    url = f"{PROSPEO_BASE}/enrich-person"
    headers = {"X-KEY": api_key, "Content-Type": "application/json"}
    body = {"data": {"person_id": person_id}, "only_verified_email": True}
    resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    if resp.status_code == 400:
        return {}  # no verified email — skip silently
    if resp.status_code == 429:
        import time; time.sleep(5)
        resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
    resp.raise_for_status()
    data = resp.json() or {}
    if data.get("error"):
        return {}
    person = data.get("person") or {}
    company = data.get("company") or {}
    email_block = person.get("email") or {}
    email = email_block.get("email") if email_block.get("status") != "UNAVAILABLE" else None
    # 1 credit if an email was returned and Prospeo billed it; 0 otherwise
    billed = 0 if data.get("free_enrichment") else (1 if email else 0)
    return {
        "email": email,
        "email_status": email_block.get("status"),
        "first_name": person.get("first_name"),
        "last_name": person.get("last_name"),
        "title": person.get("current_job_title"),
        "company_name": company.get("name") or company.get("company_name"),
        "company_domain": _norm_domain(company.get("website") or company.get("url")),
        "company_website": company.get("website") or company.get("url"),
        "company_description": company.get("description"),
        "free": data.get("free_enrichment"),
        "_credits": billed,
    }


# Legacy wrapper used by run(): swapped for the two-call workflow below.
def _prospeo_search(domain: str, api_key: str, limit: int = 25,
                    max_pages: int = 1) -> list[dict]:
    """Call Prospeo's domain-search endpoint and return a list of normalized
    person dicts shaped for downstream filtering.

    Endpoint: POST https://api.prospeo.io/domain-search
    Auth: X-KEY header
    Response: { email_list: [...], company: {...}, meta: { search_id } }

    Pagination: pass meta.search_id back in subsequent requests to get more results.
    We cap at `max_pages` to keep credit spend predictable.
    """
    url = f"{PROSPEO_BASE}/domain-search"
    headers = {"X-KEY": api_key, "Content-Type": "application/json"}
    out: list[dict] = []
    search_id: str | None = None

    for _ in range(max_pages):
        body: dict = {"company": domain, "limit": limit}
        if search_id:
            body["search_id"] = search_id

        resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
        if resp.status_code == 429:
            # naive backoff; bump if we hit this often
            import time; time.sleep(5)
            resp = requests.post(url, json=body, headers=headers, timeout=PROSPEO_TIMEOUT)
        resp.raise_for_status()
        data = resp.json() or {}

        if data.get("error"):
            print(f"  ! prospeo error for {domain}: {data}", file=sys.stderr)
            break

        company = data.get("company") or {}
        for p in data.get("email_list") or []:
            out.append({
                "email": p.get("email"),
                "first_name": p.get("first_name"),
                "last_name": p.get("last_name"),
                # domain-search uses 'position'; we map to 'title' for downstream consistency
                "title": p.get("position") or p.get("title"),
                "company_name": company.get("name") or company.get("company_name"),
                "company_domain": company.get("domain") or domain,
                "company_website": company.get("website") or company.get("url"),
                "company_description": company.get("description"),
            })

        search_id = (data.get("meta") or {}).get("search_id")
        if not search_id:
            break

    return out, credits_used


# --- filtering --------------------------------------------------------------

def is_decision_maker(title: str | None) -> bool:
    if not title:
        return False
    t = title.lower()
    return any(tok in t for tok in DECISION_MAKER_TITLES)


def rule_classify(lead: dict) -> tuple[str, str, str] | None:
    """Returns (result, method, reason) if rules can decide; else None → LLM."""
    name = (lead.get("company_name") or "").lower()
    site = _norm_domain(lead.get("company_website")) or ""
    email_dom = _email_domain(lead.get("email")) or ""

    if site in MARKETPLACE_DOMAINS or email_dom in MARKETPLACE_DOMAINS:
        return ("marketplace", "rule", "known marketplace domain")

    for tok in AGENCY_TOKENS:
        if tok in name or tok in site:
            return ("agency", "rule", f"matched agency token '{tok}'")

    # Email-domain mismatch is NOT a reliable reseller signal (many brands use
    # legacy/short email domains different from their website). Defer to LLM.
    return None


def llm_classify_batch(client: anthropic.Anthropic, system: str,
                       leads: list[dict]) -> list[tuple[str, str]]:
    """Returns [(result, reason), ...] aligned with input order."""
    results: list[tuple[str, str]] = []
    for ld in leads:
        payload = {
            "company_name": ld.get("company_name"),
            "company_website": ld.get("company_website"),
            "company_description": ld.get("company_description"),
            "title": ld.get("title"),
            "email_domain": _email_domain(ld.get("email")),
        }
        resp = client.messages.create(
            model=AGENCY_FILTER_MODEL,
            max_tokens=200,
            system=system,
            messages=[{"role": "user", "content": json.dumps(payload)}],
        )
        text = resp.content[0].text.strip() if resp.content else "{}"
        # Strip ```json ... ``` fences if present
        if text.startswith("```"):
            text = text.split("```", 2)[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.strip().rstrip("`").strip()
        try:
            parsed = json.loads(text)
            results.append((parsed.get("result", "unknown"),
                            parsed.get("reason", "")[:200]))
        except json.JSONDecodeError:
            results.append(("unknown", f"unparseable llm response: {text[:80]}"))
    return results


# --- DB write ---------------------------------------------------------------

INSERT_SQL = """
insert into prospeo_new_leads
  (email, first_name, last_name, title, company_name, company_domain,
   company_website, source_domain, source_industry, scrape_mode, prospeo_raw,
   agency_filter_result, agency_filter_method, agency_filter_reason, rejected,
   mobile, mobile_status)
values %s
on conflict (email) do nothing
"""


def write_leads(conn, rows: list[dict]) -> int:
    if not rows:
        return 0
    values = [
        (
            r["email"], r.get("first_name"), r.get("last_name"), r.get("title"),
            r.get("company_name"), r.get("company_domain"),
            r.get("company_website"), r.get("source_domain"),
            r.get("source_industry"), r.get("scrape_mode") or "domain",
            json.dumps(r.get("prospeo_raw") or {}),
            r["agency_filter_result"], r["agency_filter_method"],
            r["agency_filter_reason"], r["rejected"],
            r.get("mobile"), r.get("mobile_status"),
        )
        for r in rows
    ]
    with conn.cursor() as cur:
        execute_values(cur, INSERT_SQL, values)
    conn.commit()
    return len(rows)


CSV_COLS = ["email", "mobile", "first_name", "last_name", "title", "company_name",
            "company_website", "source_domain", "agency_filter_result"]

# XLSX adds audit columns: mobile_status, filter method/reason, and the new
# mode + source_industry so audits can slice by which path produced each row.
XLSX_COLS = CSV_COLS + ["mobile_status", "agency_filter_method", "agency_filter_reason",
                        "scrape_mode", "source_industry"]


def write_csv(rows: list[dict], path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=CSV_COLS, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(accepted: list[dict], rejected: list[dict], path: str) -> None:
    """Two-sheet workbook: 'Accepted' (brands Jam loads) + 'Rejected' (audit)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    accepted_fill = PatternFill("solid", fgColor="1A7F37")
    rejected_fill = PatternFill("solid", fgColor="CF222E")

    def fill_sheet(ws, rows, fill):
        ws.append(XLSX_COLS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
        for r in rows:
            ws.append([r.get(c, "") for c in XLSX_COLS])
        # Autosize columns (rough)
        for col_idx, col_name in enumerate(XLSX_COLS, start=1):
            max_len = max((len(str(r.get(col_name, ""))) for r in rows), default=0)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
                min(max(len(col_name), max_len) + 2, 50)

    ws_a = wb.active
    ws_a.title = "Accepted"
    fill_sheet(ws_a, accepted, accepted_fill)

    ws_r = wb.create_sheet("Rejected")
    fill_sheet(ws_r, rejected, rejected_fill)

    wb.save(path)


# --- main -------------------------------------------------------------------

def run(mode: str = "domain", *,
        domains_csv: str | None = None, limit: int | None = None,
        target_leads: int | None = None,
        country: list[str] | None = None,
        dry_run: bool = False, skip_llm: bool = False,
        with_mobile: bool = False,
        max_credits: int | None = None) -> None:
    """Dispatch to the appropriate scraper based on mode.

    mode='domain'   -> _run_domain (today's behavior — query Prospeo per
                       inclusion-list domain, batch in groups of 15,
                       dedup by domain + email).
    mode='category' -> _run_category (Step 4 — query Prospeo by industry,
                       paginate per industry with state in
                       category_scrape_state, dedup by email only).

    target_leads + country are only used by category mode.
    domains_csv + limit are only used by domain mode.
    """
    load_dotenv()
    api_key = os.environ.get("PROSPEO_API_KEY", "").strip()
    if not api_key and not dry_run:
        sys.exit("PROSPEO_API_KEY not set in .env")

    conn = connect()
    try:
        if mode == "domain":
            _run_domain(conn, api_key,
                        domains_csv=domains_csv, limit=limit,
                        dry_run=dry_run, skip_llm=skip_llm,
                        with_mobile=with_mobile, max_credits=max_credits)
        elif mode == "category":
            _run_category(conn, api_key,
                          target_leads=target_leads, country=country,
                          dry_run=dry_run, skip_llm=skip_llm,
                          with_mobile=with_mobile, max_credits=max_credits)
        else:
            sys.exit(f"Unknown mode: {mode!r}. Expected 'domain' or 'category'.")
    finally:
        conn.close()


def _run_domain(conn, api_key: str, *,
                domains_csv: str | None, limit: int | None,
                dry_run: bool, skip_llm: bool, with_mobile: bool,
                max_credits: int | None) -> None:
    """Domain mode: query Prospeo per inclusion-list domain, batched in 15s.

    Pre-Step-3 this was the body of run(). Lifted as-is into its own
    function; behavior identical. conn + api_key are passed in by the
    new run() dispatcher (which also owns the connection lifecycle).
    """
    try:
        if domains_csv:
            domains = load_domains_from_csv(domains_csv)
        else:
            domains = load_domains_from_db(conn, limit=limit)

        if limit:
            domains = domains[:limit]

        if not domains:
            sys.exit("No domains to scrape. Pass --domains <csv> or populate domain_inclusion_list.")

        skip_domains = fetch_domains_with_decision_maker(conn)
        already_scraped = fetch_recently_scraped_domains(conn)
        existing_emails = fetch_existing_emails(conn)

        initial_count = len(domains)
        domains_set = set(domains)
        skipped_decision_maker = len(skip_domains & domains_set)
        skipped_already_scraped = len(already_scraped & domains_set)

        print(f"domains: {initial_count}")
        print(f"  skipping {skipped_decision_maker} (already have a decision-maker)")
        print(f"  skipping {skipped_already_scraped} (already scraped in the last 30 days)")
        domains = [d for d in domains
                   if d not in skip_domains and d not in already_scraped]
        print(f"to query: {len(domains)}")

        if dry_run:
            print("Dry run — first 5 domains:")
            for d in domains[:5]:
                print(f"  {d}")
            return

        accepted: list[dict] = []
        rejected: list[dict] = []
        searched = 0
        enriched_with_email = 0
        enriched_without_email = 0
        credits_spent = 0
        aborted_reason: str | None = None

        # Anthropic client lazily so the script still runs offline if no key
        llm_client = None
        llm_system = None

        # Step 1: batched search-person across all domains
        # Step 2: enrich-person per person_id
        # Step 3: rule + LLM agency filter PER BATCH (so checkpoint is complete)
        # Step 4: write batch to DB immediately (per-batch checkpoint)
        batches = [domains[i:i + PROSPEO_BATCH_DOMAINS]
                   for i in range(0, len(domains), PROSPEO_BATCH_DOMAINS)]

        for bi, batch in enumerate(batches, 1):
            # Budget cap check before any API call in this batch
            if max_credits is not None and credits_spent >= max_credits:
                aborted_reason = f"budget cap hit ({credits_spent}/{max_credits} credits)"
                break

            try:
                hits, search_credits = _search_people(batch, api_key)
            except requests.RequestException as e:
                print(f"  ! batch {bi}/{len(batches)} search-person failed: {e}", file=sys.stderr)
                continue
            credits_spent += search_credits
            searched += len(hits)

            batch_accepted: list[dict] = []
            batch_rejected: list[dict] = []
            batch_grey: list[dict] = []

            for h in hits:
                # Budget cap check before enrich
                if max_credits is not None and credits_spent >= max_credits:
                    aborted_reason = f"budget cap hit ({credits_spent}/{max_credits} credits)"
                    break

                pid = h.get("person_id")
                if not pid:
                    continue
                try:
                    en = _enrich_person(pid, api_key)
                except requests.RequestException as e:
                    print(f"  ! enrich-person {pid}: {e}", file=sys.stderr)
                    continue
                credits_spent += en.get("_credits", 0)

                email = (en.get("email") or "").lower().strip()
                if not email:
                    enriched_without_email += 1
                    continue
                if email in existing_emails:
                    continue
                enriched_with_email += 1
                existing_emails.add(email)

                # Map enriched company back to source domain (best effort)
                source_dom = _norm_domain(en.get("company_website")) or _norm_domain(h.get("company_website")) or ""
                source_dom = source_dom if source_dom in set(batch) else (next(
                    (d for d in batch if d in (source_dom or "")), source_dom))

                lead = {
                    "email": email,
                    "first_name": en.get("first_name") or h.get("first_name"),
                    "last_name": en.get("last_name") or h.get("last_name"),
                    "title": en.get("title") or h.get("title"),
                    "company_name": en.get("company_name") or h.get("company_name"),
                    "company_domain": en.get("company_domain"),
                    "company_website": en.get("company_website") or h.get("company_website"),
                    "company_description": en.get("company_description") or h.get("company_description"),
                    "source_domain": source_dom,
                    "person_id": pid,
                    "mobile": None,
                    "mobile_status": None,
                    "prospeo_raw": {"search": h, "enrich": {k: v for k, v in en.items()
                                                              if k not in ("company_description", "_credits")}},
                }

                rule = rule_classify(lead)
                if rule:
                    result, method, reason = rule
                    lead.update(agency_filter_result=result,
                                agency_filter_method=method,
                                agency_filter_reason=reason,
                                rejected=result != "brand")
                    (batch_rejected if lead["rejected"] else batch_accepted).append(lead)
                else:
                    batch_grey.append(lead)

            # Per-batch LLM grey-zone classification
            if batch_grey:
                if not skip_llm:
                    if llm_client is None:
                        llm_client = anthropic.Anthropic()
                        llm_system = AGENCY_FILTER_PROMPT.read_text(encoding="utf-8")
                    outcomes = llm_classify_batch(llm_client, llm_system, batch_grey)
                    for lead, (result, reason) in zip(batch_grey, outcomes):
                        lead.update(agency_filter_result=result,
                                    agency_filter_method="llm",
                                    agency_filter_reason=reason,
                                    rejected=result != "brand")
                        (batch_rejected if lead["rejected"] else batch_accepted).append(lead)
                else:
                    for lead in batch_grey:
                        lead.update(agency_filter_result="unknown",
                                    agency_filter_method="none",
                                    agency_filter_reason="llm skipped",
                                    rejected=False)
                        batch_accepted.append(lead)

            # CHECKPOINT: persist this batch's leads + mark domains as scraped
            if batch_accepted or batch_rejected:
                write_leads(conn, batch_accepted + batch_rejected)
                accepted.extend(batch_accepted)
                rejected.extend(batch_rejected)
            # Always mark the domains as scraped, even if zero leads found
            # (otherwise we'd re-query empty domains forever)
            mark_domains_scraped(conn, batch)

            # Live progress line
            print(f"  [batch {bi}/{len(batches)}] "
                  f"credits: {credits_spent}"
                  f"{f'/{max_credits}' if max_credits else ''} | "
                  f"accepted: {len(accepted)} | rejected: {len(rejected)}")

            if aborted_reason:
                break

        if aborted_reason:
            print(f"\n!!! Run aborted: {aborted_reason}", file=sys.stderr)

        credits_spent, _mobile_found, csv_path, xlsx_path = _finalize_and_export(
            conn, api_key,
            accepted=accepted, rejected=rejected,
            credits_spent=credits_spent, aborted_reason=aborted_reason,
            with_mobile=with_mobile, max_credits=max_credits,
        )

        print(f"\n=== summary ===")
        if aborted_reason:
            print(f"  ABORTED:              {aborted_reason}")
        print(f"  domains queried:      {len(domains)}")
        print(f"  decision-makers found: {searched}")
        print(f"    with email:         {enriched_with_email}")
        print(f"    no email available: {enriched_without_email}")
        print(f"  accepted (brand):     {len(accepted)}")
        print(f"  rejected:             {len(rejected)}")
        print(f"  total credits spent:  {credits_spent}  (~${credits_spent * 0.02:.2f})")
        print(f"  CSV (accepted):       {csv_path}")
        print(f"  XLSX (both sheets):   {xlsx_path}")
    finally:
        # conn lifecycle (close) is now owned by the run() dispatcher.
        pass


def _finalize_and_export(conn, api_key: str, *,
                          accepted: list[dict], rejected: list[dict],
                          credits_spent: int, aborted_reason: str | None,
                          with_mobile: bool, max_credits: int | None
                          ) -> tuple[int, int, str, str]:
    """Shared post-pass for both modes.

    Steps:
      1. Optional mobile enrichment on accepted leads (10 credits per verified
         mobile; skipped if run was aborted).
      2. Write CSV (accepted only — what Jam loads into Instantly).
      3. Write XLSX (Accepted + Rejected sheets — for audit).

    Returns (credits_spent_after_mobile, mobile_found, csv_path, xlsx_path).
    Caller prints its own mode-specific summary using these values.
    """
    mobile_found = 0
    if with_mobile and accepted and not aborted_reason:
        print(f"\nEnriching mobile for {len(accepted)} accepted leads (~10 credits each)...")
        for lead in accepted:
            if max_credits is not None and credits_spent >= max_credits:
                print(f"  ! budget cap hit during mobile enrichment "
                      f"({credits_spent}/{max_credits})")
                break
            pid = lead.get("person_id")
            if not pid:
                continue
            try:
                m = _enrich_mobile(pid, api_key)
            except requests.RequestException as e:
                print(f"  ! enrich-mobile {pid}: {e}", file=sys.stderr)
                continue
            credits_spent += m.get("_credits", 0)
            if m.get("mobile"):
                lead["mobile"] = m["mobile"]
                lead["mobile_status"] = m.get("mobile_status")
                with conn.cursor() as cur:
                    cur.execute(
                        "update prospeo_new_leads set mobile=%s, mobile_status=%s where email=%s",
                        (m["mobile"], m.get("mobile_status"), lead["email"]),
                    )
                conn.commit()
                mobile_found += 1
        print(f"  mobile found: {mobile_found}/{len(accepted)} (credits now: {credits_spent})")

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    os.makedirs("exports", exist_ok=True)
    csv_path = f"exports/prospeo_new_leads_{stamp}.csv"
    xlsx_path = f"exports/prospeo_new_leads_{stamp}.xlsx"
    write_csv(accepted, csv_path)
    write_xlsx(accepted, rejected, xlsx_path)
    return credits_spent, mobile_found, csv_path, xlsx_path


def _run_category(conn, api_key: str, *,
                   target_leads: int | None,
                   country: list[str] | None,
                   dry_run: bool, skip_llm: bool, with_mobile: bool,
                   max_credits: int | None) -> None:
    """Category mode: query Prospeo by industry, paginate per industry.

    For each industry in PROSPEO_INDUSTRIES, round-robin one page at a time,
    enriching results and applying the same rule+LLM filter as domain mode.
    Pagination cursor is persisted in category_scrape_state so subsequent
    runs resume where the last one left off.

    Stops when any of these is true:
      - accepted-lead count reaches target_leads (default: no target → keep
        going until budget or all industries exhausted)
      - credits_spent reaches max_credits
      - all industries are exhausted (last_page_consumed >= total_pages)
    """
    countries = list(country or [])

    if dry_run:
        print(f"Dry run — category mode")
        print(f"  industries: {len(PROSPEO_INDUSTRIES)}")
        for ind in PROSPEO_INDUSTRIES:
            print(f"    - {ind}")
        print(f"  titles:     {len(PROSPEO_TITLES)} (two-tier: owner + marketing/e-com)")
        print(f"  countries:  {countries or '(none — global)'}")
        print(f"  target_leads: {target_leads or '(unlimited)'}")
        print(f"  max_credits:  {max_credits or '(uncapped)'}")
        return

    state = fetch_category_state(conn)
    existing_emails = fetch_existing_emails(conn)

    # Build the round-robin queue: industries not yet exhausted.
    queue: list[str] = []
    for ind in PROSPEO_INDUSTRIES:
        s = state.get(ind, {})
        if s.get("exhausted"):
            print(f"  [skip] {ind!r} exhausted "
                  f"(page {s.get('last_page_consumed')}/{s.get('total_pages')})")
            continue
        queue.append(ind)

    if not queue:
        print("All industries exhausted. Run "
              "`update category_scrape_state set exhausted=false, last_page_consumed=0` "
              "to re-scan from the top.")
        return

    print(f"category mode — {len(queue)} industries to scan, "
          f"countries={countries or '(global)'}")

    accepted: list[dict] = []
    rejected: list[dict] = []
    credits_spent = 0
    enriched_with_email = 0
    enriched_without_email = 0
    aborted_reason: str | None = None
    llm_client = None
    llm_system = None
    pages_credits_added: dict[str, int] = {ind: 0 for ind in queue}

    # Round-robin: one page per industry per cycle. Stop when budget/target hit
    # or when all remaining industries became exhausted this run.
    while queue:
        if target_leads is not None and len(accepted) >= target_leads:
            aborted_reason = f"target_leads reached ({len(accepted)}/{target_leads})"
            break
        if max_credits is not None and credits_spent >= max_credits:
            aborted_reason = f"budget cap hit ({credits_spent}/{max_credits} credits)"
            break

        next_queue: list[str] = []
        for ind in queue:
            if target_leads is not None and len(accepted) >= target_leads:
                aborted_reason = f"target_leads reached ({len(accepted)}/{target_leads})"
                break
            if max_credits is not None and credits_spent >= max_credits:
                aborted_reason = f"budget cap hit ({credits_spent}/{max_credits} credits)"
                break

            s = state.get(ind, {})
            next_page = (s.get("last_page_consumed") or 0) + 1

            try:
                results, total_pages, page_credits = _search_people_by_industry(
                    ind, countries, next_page, api_key)
            except InsufficientCreditsError as e:
                # Hard abort — don't retry, don't touch state.
                print(f"\n!!! {e}", file=sys.stderr)
                aborted_reason = "Prospeo INSUFFICIENT_CREDITS — top up and re-run"
                break
            except requests.RequestException as e:
                print(f"  ! search-person {ind!r} page {next_page}: {e}", file=sys.stderr)
                next_queue.append(ind)  # try again next cycle
                continue

            credits_spent += page_credits
            pages_credits_added[ind] = pages_credits_added.get(ind, 0) + page_credits

            batch_accepted: list[dict] = []
            batch_rejected: list[dict] = []
            batch_grey: list[dict] = []

            for h in results:
                if max_credits is not None and credits_spent >= max_credits:
                    aborted_reason = f"budget cap hit ({credits_spent}/{max_credits} credits)"
                    break

                pid = h.get("person_id")
                if not pid:
                    continue
                try:
                    en = _enrich_person(pid, api_key)
                except requests.RequestException as e:
                    print(f"  ! enrich-person {pid}: {e}", file=sys.stderr)
                    continue
                credits_spent += en.get("_credits", 0)
                pages_credits_added[ind] = pages_credits_added.get(ind, 0) + en.get("_credits", 0)

                email = (en.get("email") or "").lower().strip()
                if not email:
                    enriched_without_email += 1
                    continue
                if email in existing_emails:
                    continue
                enriched_with_email += 1
                existing_emails.add(email)

                lead = {
                    "email": email,
                    "first_name": en.get("first_name") or h.get("first_name"),
                    "last_name": en.get("last_name") or h.get("last_name"),
                    "title": en.get("title") or h.get("title"),
                    "company_name": en.get("company_name") or h.get("company_name"),
                    "company_domain": en.get("company_domain"),
                    "company_website": en.get("company_website") or h.get("company_website"),
                    "company_description": en.get("company_description") or h.get("company_description"),
                    "source_domain": None,
                    "source_industry": ind,
                    "scrape_mode": "category",
                    "person_id": pid,
                    "mobile": None,
                    "mobile_status": None,
                    "prospeo_raw": {"search": h,
                                    "enrich": {k: v for k, v in en.items()
                                               if k not in ("company_description", "_credits")}},
                }

                rule = rule_classify(lead)
                if rule:
                    result, method, reason = rule
                    lead.update(agency_filter_result=result,
                                agency_filter_method=method,
                                agency_filter_reason=reason,
                                rejected=result != "brand")
                    (batch_rejected if lead["rejected"] else batch_accepted).append(lead)
                else:
                    batch_grey.append(lead)

            # LLM grey-zone for this page
            if batch_grey:
                if not skip_llm:
                    if llm_client is None:
                        llm_client = anthropic.Anthropic()
                        llm_system = AGENCY_FILTER_PROMPT.read_text(encoding="utf-8")
                    outcomes = llm_classify_batch(llm_client, llm_system, batch_grey)
                    for lead, (result, reason) in zip(batch_grey, outcomes):
                        lead.update(agency_filter_result=result,
                                    agency_filter_method="llm",
                                    agency_filter_reason=reason,
                                    rejected=result != "brand")
                        (batch_rejected if lead["rejected"] else batch_accepted).append(lead)
                else:
                    for lead in batch_grey:
                        lead.update(agency_filter_result="unknown",
                                    agency_filter_method="none",
                                    agency_filter_reason="llm skipped",
                                    rejected=False)
                        batch_accepted.append(lead)

            # CHECKPOINT: persist this page's leads + update state row
            if batch_accepted or batch_rejected:
                write_leads(conn, batch_accepted + batch_rejected)
                accepted.extend(batch_accepted)
                rejected.extend(batch_rejected)

            exhausted = next_page >= (total_pages or 1)
            upsert_category_state(
                conn, industry=ind, countries=countries,
                last_page_consumed=next_page, total_pages=total_pages or None,
                exhausted=exhausted, credits_added=pages_credits_added.get(ind, 0),
            )
            pages_credits_added[ind] = 0  # reset, already persisted
            state[ind] = {
                "countries": countries, "last_page_consumed": next_page,
                "total_pages": total_pages, "exhausted": exhausted,
                "last_scraped_at": None, "total_credits_spent": 0,
            }

            print(f"  [{ind}] page {next_page}/{total_pages or '?'}: "
                  f"+{len(batch_accepted)} accepted / +{len(batch_rejected)} rejected "
                  f"(credits: {credits_spent}{f'/{max_credits}' if max_credits else ''})")

            if not exhausted:
                next_queue.append(ind)

            if aborted_reason:
                break

        if aborted_reason:
            break
        queue = next_queue

    if aborted_reason:
        print(f"\n!!! Run aborted: {aborted_reason}", file=sys.stderr)

    credits_spent, _mobile_found, csv_path, xlsx_path = _finalize_and_export(
        conn, api_key,
        accepted=accepted, rejected=rejected,
        credits_spent=credits_spent, aborted_reason=aborted_reason,
        with_mobile=with_mobile, max_credits=max_credits,
    )

    # Per-industry pages-consumed summary
    print(f"\n=== summary ===")
    if aborted_reason:
        print(f"  ABORTED:                {aborted_reason}")
    print(f"  countries:              {countries or '(global)'}")
    print(f"  industries scanned:     {len(PROSPEO_INDUSTRIES)}")
    exhausted_count = sum(1 for ind in PROSPEO_INDUSTRIES
                          if state.get(ind, {}).get("exhausted"))
    print(f"    exhausted (cumulative): {exhausted_count}")
    print(f"  decision-makers found:  {enriched_with_email + enriched_without_email}")
    print(f"    with email:           {enriched_with_email}")
    print(f"    no email available:   {enriched_without_email}")
    print(f"  accepted (brand):       {len(accepted)}")
    print(f"  rejected:               {len(rejected)}")
    print(f"  total credits spent:    {credits_spent}  (~${credits_spent * 0.02:.2f})")
    print(f"  CSV (accepted):         {csv_path}")
    print(f"  XLSX (both sheets):     {xlsx_path}")


def export_all_leads(out_dir: str = "exports") -> tuple[str, str]:
    """Dump the full prospeo_new_leads table into a fresh CSV + XLSX.

    The CSV contains only accepted brands (what Jam loads into Instantly).
    The XLSX has two sheets — Accepted + Rejected (with filter reasons) —
    for review/audit.

    Returns the (csv_path, xlsx_path) pair.
    """
    conn = connect()
    accepted: list[dict] = []
    rejected: list[dict] = []
    try:
        with conn.cursor() as cur:
            cur.execute("""
              select email, mobile, first_name, last_name, title, company_name,
                     company_website, source_domain, agency_filter_result,
                     mobile_status, agency_filter_method, agency_filter_reason, rejected
              from prospeo_new_leads
              order by rejected, scraped_at desc
            """)
            for r in cur.fetchall():
                lead = {
                    "email": r[0], "mobile": r[1],
                    "first_name": r[2], "last_name": r[3], "title": r[4],
                    "company_name": r[5], "company_website": r[6],
                    "source_domain": r[7], "agency_filter_result": r[8],
                    "mobile_status": r[9], "agency_filter_method": r[10],
                    "agency_filter_reason": r[11],
                }
                if r[12]:
                    rejected.append(lead)
                else:
                    accepted.append(lead)
    finally:
        conn.close()

    os.makedirs(out_dir, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    csv_path = f"{out_dir}/cumulative_{stamp}.csv"
    xlsx_path = f"{out_dir}/cumulative_{stamp}.xlsx"
    write_csv(accepted, csv_path)
    write_xlsx(accepted, rejected, xlsx_path)

    print(f"Exported {len(accepted)} accepted brand(s) + {len(rejected)} rejected:")
    print(f"  CSV  : {csv_path}")
    print(f"  XLSX : {xlsx_path}")
    return csv_path, xlsx_path


def enrich_mobile_for_accepted(limit: int | None = None,
                                dry_run: bool = False) -> None:
    """Catch-up command: enrich mobile for every accepted lead in
    prospeo_new_leads that doesn't have one yet.

    Cost: 10 credits per verified mobile found; free if not available."""
    load_dotenv()
    api_key = os.environ.get("PROSPEO_API_KEY", "").strip()
    if not api_key and not dry_run:
        sys.exit("PROSPEO_API_KEY not set in .env")

    conn = connect()
    try:
        with conn.cursor() as cur:
            sql = """
              select id, email, prospeo_raw
              from prospeo_new_leads
              where rejected = false and mobile is null
              order by scraped_at desc
            """
            if limit:
                sql += f" limit {int(limit)}"
            cur.execute(sql)
            rows = cur.fetchall()

        if not rows:
            print("No accepted leads pending mobile enrichment.")
            return

        print(f"{len(rows)} accepted lead(s) pending mobile enrichment.")
        print(f"Estimated cost: up to {len(rows) * 10} credits "
              f"(only charged for verified mobiles actually returned).")
        if dry_run:
            print("Dry run — exiting before any API call.")
            return

        found = 0
        for row_id, email, raw in rows:
            pid = ((raw or {}).get("search") or {}).get("person_id") if isinstance(raw, dict) else None
            if not pid:
                print(f"  ! {email}: no person_id stored, skipping")
                continue
            try:
                m = _enrich_mobile(pid, api_key)
            except requests.RequestException as e:
                print(f"  ! {email}: {e}", file=sys.stderr)
                continue
            if m.get("mobile"):
                with conn.cursor() as cur:
                    cur.execute(
                        "update prospeo_new_leads "
                        "set mobile=%s, mobile_status=%s where id=%s",
                        (m["mobile"], m.get("mobile_status"), row_id),
                    )
                found += 1
                print(f"  + {email}: {m['mobile']}")
        conn.commit()
        print(f"\nMobile enriched: {found}/{len(rows)}")
    finally:
        conn.close()


def main(domains_csv: str | None = None, limit: int | None = None,
         dry_run: bool = False, skip_llm: bool = False,
         with_mobile: bool = False, max_credits: int | None = None,
         mode: str = "domain",
         target_leads: int | None = None,
         country: list[str] | None = None) -> None:
    """Entry point called by run.py.

    Keeps the original positional/keyword arg surface intact for backward
    compatibility. New mode-related kwargs default to domain-mode behavior.
    """
    run(mode=mode,
        domains_csv=domains_csv, limit=limit,
        target_leads=target_leads, country=country,
        dry_run=dry_run, skip_llm=skip_llm,
        with_mobile=with_mobile, max_credits=max_credits)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--domains", help="CSV of domains; defaults to domain_inclusion_list table")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--skip-llm", action="store_true")
    ap.add_argument("--with-mobile", action="store_true",
                    help="Enrich accepted leads with mobile numbers (~10 credits each)")
    ap.add_argument("--max-credits", type=int, default=None,
                    help="Hard budget cap. Abort before any call that would push spend past this.")
    args = ap.parse_args()
    main(args.domains, args.limit, args.dry_run, args.skip_llm,
         args.with_mobile, args.max_credits)
