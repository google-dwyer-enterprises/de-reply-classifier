"""Upsert Apollo enrichment data into lead_contacts.

Re-runnable: matches on lead_email (= first address from Apollo's "Emails"
column, lowercased+trimmed). New emails are inserted; existing rows have
every Apollo column overwritten and updated_at bumped. Never touches the
`leads` table.

CLI: python run.py upload-leads <file>
Accepts .csv or .xlsx.
"""

from __future__ import annotations

import csv
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from supabase import create_client


HEADER_MAP: dict[str, str] = {
    "Apollo Company Name": "apollo_company_name",
    "Domains Match": "domains_match",
    "Lead List Source": "lead_list_source",
    "Company Name": "company_name",
    "Full Name": "full_name",
    "First Name": "first_name",
    "Last Name": "last_name",
    "Emails": "emails",
    "Website": "website",
    "Title": "title",
    "Seniority": "seniority",
    "Departments": "departments",
    "# Employees": "num_employees",
    "Industry": "industry",
    "Keywords": "keywords",
    "Person Linkedin Url": "person_linkedin_url",
    "Company Linkedin Url": "company_linkedin_url",
    "Facebook Url": "facebook_url",
    "Twitter Url": "twitter_url",
    "City": "city",
    "State": "state",
    "Country": "country",
    "Company Phone Number": "company_phone_number",
    "Company Address": "company_address",
    "Company City": "company_city",
    "Company State": "company_state",
    "Company Country": "company_country",
    "SEO Description": "seo_description",
    "Technologies": "technologies",
    "Total Funding": "total_funding",
    "Annual Revenue/Monthly Revenue": "annual_revenue",
    "Monthly Revenue": "monthly_revenue",
    "Amazon Storefront": "amazon_storefront",
}

CHUNK_SIZE = 500
MAX_RETRIES = 5
RETRY_BASE_DELAY = 2.0


def upsert_with_retry(supabase, chunk: list[dict]) -> None:
    last_exc: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            supabase.table("lead_contacts").upsert(
                chunk, on_conflict="lead_email"
            ).execute()
            return
        except Exception as exc:
            last_exc = exc
            if attempt == MAX_RETRIES:
                break
            delay = RETRY_BASE_DELAY * (2 ** (attempt - 1))
            print(f"    retry {attempt}/{MAX_RETRIES - 1} after error: "
                  f"{type(exc).__name__}: {exc} (sleeping {delay:.1f}s)")
            time.sleep(delay)
    raise last_exc  # type: ignore[misc]


def _resolve_path(raw: str) -> Path:
    p = Path(raw)
    if p.exists():
        return p
    for ext in (".csv", ".xlsx"):
        candidate = Path(f"{raw}{ext}")
        if candidate.exists():
            return candidate
    sys.exit(f"File not found: {raw} (also tried .csv, .xlsx)")


def _iter_csv_rows(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield row


def _iter_xlsx_rows(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows)]
    for row in rows:
        yield {h: ("" if v is None else str(v)) for h, v in zip(headers, row)}


def iter_rows(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _iter_csv_rows(path)
    if suffix == ".xlsx":
        return _iter_xlsx_rows(path)
    sys.exit(f"Unsupported file type: {suffix} (expected .csv or .xlsx)")


def first_email(raw: str | None) -> str | None:
    if not raw:
        return None
    first = raw.split(",")[0].strip().lower()
    return first or None


def to_db_row(src: dict, now_iso: str) -> dict | None:
    email = first_email(src.get("Emails"))
    if not email or "@" not in email:
        return None
    out: dict = {"lead_email": email, "updated_at": now_iso}
    for header, col in HEADER_MAP.items():
        val = src.get(header)
        if val is None:
            continue
        val = str(val).strip()
        out[col] = val if val else None
    return out


def chunked(items: list, size: int):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def main(file_arg: str) -> None:
    load_dotenv()
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not url or not key:
        sys.exit("SUPABASE_URL and SUPABASE_KEY must be set in .env")

    path = _resolve_path(file_arg)
    print(f"Reading {path}")

    now_iso = datetime.now(timezone.utc).isoformat()

    rows_by_email: dict[str, dict] = {}
    skipped = 0
    duplicates_in_file = 0

    for src in iter_rows(path):
        db_row = to_db_row(src, now_iso)
        if db_row is None:
            skipped += 1
            continue
        email = db_row["lead_email"]
        existing = rows_by_email.get(email)
        if existing is None:
            rows_by_email[email] = db_row
        else:
            duplicates_in_file += 1
            new_src = db_row.get("lead_list_source")
            if new_src:
                cur = existing.get("lead_list_source") or ""
                parts = [p.strip() for p in cur.split(";") if p.strip()]
                if new_src not in parts:
                    parts.append(new_src)
                existing["lead_list_source"] = "; ".join(parts)
            for col, val in db_row.items():
                if val and not existing.get(col):
                    existing[col] = val

    rows = list(rows_by_email.values())
    total = len(rows)
    print(f"Parsed: {total} unique rows, {skipped} skipped (no email), "
          f"{duplicates_in_file} duplicate occurrences merged")

    if not rows:
        print("Nothing to upload.")
        return

    supabase = create_client(url, key)
    upserted = 0
    for i, chunk in enumerate(chunked(rows, CHUNK_SIZE), start=1):
        upsert_with_retry(supabase, chunk)
        upserted += len(chunk)
        print(f"  chunk {i}: {upserted}/{total} upserted")

    print(f"Done. {upserted} rows upserted into lead_contacts.")

    print("Refreshing lead_status materialized view...")
    from db import refresh_lead_status
    refresh_lead_status()
    print("Refreshed.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("Usage: python lead_contacts_upload.py <file.csv|.xlsx>")
    main(sys.argv[1])
