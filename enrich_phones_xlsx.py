"""Enrich missing phone numbers in an XLSX sheet via BetterContact.

Generic — works on any XLSX as long as the target sheet has a "name" column
(any of: Contact Name / Full Name / Name / first+last) and an email column.
A "phone" column is optional; if present, rows with a non-empty phone are
skipped (no re-enrichment).

The script writes the enriched phone into a NEW column (default header:
"Phone (BC enriched)") so the original phone column stays untouched. Saves
to a new XLSX (input filename + "_enriched_<stamp>.xlsx") by default so the
source file is never overwritten.

Pricing reminder: BetterContact charges 10 credits per VERIFIED phone found,
0 for not_found. A --budget-credits cap prevents overshoot.

CLI:
  python enrich_phones_xlsx.py \\
      --input "original_data/Interested leads all Client.xlsx" \\
      --sheet Sheet4 \\
      --budget-credits 100

Async behavior — BC's enrichment endpoint is async:
  POST https://app.bettercontact.rocks/api/v2/async   → request_id
  GET  https://app.bettercontact.rocks/api/v2/async/{id}  → poll until done
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

BC_BASE = "https://app.bettercontact.rocks/api/v2"
BC_BATCH_SIZE = 100              # API max per request
BC_POLL_INTERVAL_S = 10
BC_POLL_TIMEOUT_S = 600
BC_REQUEST_TIMEOUT_S = 30

# Column-name candidates we try (case-insensitive, normalized to lowercase + stripped).
NAME_CANDIDATES = ("contact name", "full name", "name", "lead name", "person name")
FIRST_NAME_CANDIDATES = ("first name", "first_name", "firstname", "given name")
LAST_NAME_CANDIDATES = ("last name", "last_name", "lastname", "family name", "surname")
EMAIL_CANDIDATES = ("email", "email address", "e-mail", "work email")
COMPANY_CANDIDATES = ("company name", "company", "organization", "employer", "business")
PHONE_CANDIDATES = ("phone", "phone number", "mobile", "telephone")

OUTPUT_PHONE_HEADER_DEFAULT = "Phone (BC enriched)"


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def find_col(header: list, candidates: tuple[str, ...]) -> int | None:
    """Return 1-indexed column number whose header matches any candidate (case-insens)."""
    for idx, h in enumerate(header, 1):
        if _norm(h) in candidates:
            return idx
    return None


def split_name(full: str) -> tuple[str, str]:
    parts = full.strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def email_to_domain(email: str | None) -> str | None:
    if not email or "@" not in email:
        return None
    return email.rsplit("@", 1)[1].strip().lower()


# ---------------------------------------------------------------------------
# BC enrichment async API
# ---------------------------------------------------------------------------

def bc_submit(payload: dict, api_key: str) -> str:
    headers = {"X-API-Key": api_key, "Content-Type": "application/json"}
    r = requests.post(f"{BC_BASE}/async", json=payload, headers=headers,
                      timeout=BC_REQUEST_TIMEOUT_S)
    if r.status_code not in (200, 201, 202):
        raise RuntimeError(f"BC enrich submit {r.status_code}: {r.text[:300]}")
    return r.json().get("id") or r.json().get("request_id")


def bc_poll(request_id: str, api_key: str, timeout_s: int = BC_POLL_TIMEOUT_S) -> dict:
    headers = {"X-API-Key": api_key}
    url = f"{BC_BASE}/async/{request_id}"
    start = time.time()
    while time.time() - start < timeout_s:
        time.sleep(BC_POLL_INTERVAL_S)
        try:
            r = requests.get(url, headers=headers, timeout=BC_REQUEST_TIMEOUT_S)
        except Exception:
            continue
        if r.status_code != 200 and r.status_code != 202:
            continue
        data = r.json() or {}
        status = (data.get("status") or "").lower()
        if status in ("terminated", "completed", "done", "finished"):
            return data
    raise RuntimeError(f"BC enrich poll {request_id} timeout after {timeout_s}s")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", required=True, help="Input XLSX path")
    p.add_argument("--sheet", required=True, help="Sheet name to enrich")
    p.add_argument("--output", default=None,
                   help="Output XLSX path (default: <input>_enriched_<stamp>.xlsx)")
    p.add_argument("--phone-header", default=OUTPUT_PHONE_HEADER_DEFAULT,
                   help=f"Column header for the new enriched-phone column "
                        f"(default: {OUTPUT_PHONE_HEADER_DEFAULT!r})")
    p.add_argument("--budget-credits", type=int, default=None,
                   help="Hard cap on credits to spend (10/found phone). "
                        "Default: unlimited; rely on BC's own balance check.")
    p.add_argument("--dry-run", action="store_true",
                   help="Detect columns + count enrichable rows, do NOT call BC")
    p.add_argument("--limit", type=int, default=None,
                   help="Process at most N enrichable rows (smoke test)")
    args = p.parse_args()

    load_dotenv()
    api_key = (os.environ.get("BETTERCONTACT_API_KEY") or "").strip()
    if not api_key and not args.dry_run:
        sys.exit("BETTERCONTACT_API_KEY not set in .env")

    # Output path default
    if not args.output:
        in_path = Path(args.input)
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        args.output = str(in_path.with_name(f"{in_path.stem}_enriched_{stamp}.xlsx"))

    # Load workbook (NOT read_only — we need to edit + save)
    wb = load_workbook(args.input)
    if args.sheet not in wb.sheetnames:
        sys.exit(f"Sheet {args.sheet!r} not found. Available: {wb.sheetnames}")
    ws = wb[args.sheet]

    header = [c.value for c in ws[1]]
    name_col = find_col(header, NAME_CANDIDATES)
    first_col = find_col(header, FIRST_NAME_CANDIDATES)
    last_col = find_col(header, LAST_NAME_CANDIDATES)
    email_col = find_col(header, EMAIL_CANDIDATES)
    company_col = find_col(header, COMPANY_CANDIDATES)
    phone_col = find_col(header, PHONE_CANDIDATES)

    if email_col is None:
        sys.exit("No email column found. Headers: " + repr(header))
    if name_col is None and (first_col is None or last_col is None):
        sys.exit("No name column found (need 'Contact Name' OR 'First/Last Name'). "
                 "Headers: " + repr(header))

    print(f"Detected columns:")
    print(f"  name      = col {name_col}  ({header[name_col-1]!r})" if name_col else f"  name      = (none)")
    print(f"  first     = col {first_col} ({header[first_col-1]!r})" if first_col else f"  first     = (none)")
    print(f"  last      = col {last_col}  ({header[last_col-1]!r})" if last_col else f"  last      = (none)")
    print(f"  email     = col {email_col} ({header[email_col-1]!r})")
    print(f"  company   = col {company_col} ({header[company_col-1]!r})" if company_col else f"  company   = (none)")
    print(f"  phone     = col {phone_col}  ({header[phone_col-1]!r})" if phone_col else f"  phone     = (none — every row is a candidate)")

    # Output column. If the sheet already has a phone column, write enriched
    # phones into THAT column's empty cells (no duplicate column). If no phone
    # column exists, append a fresh one with the --phone-header label.
    if phone_col is not None:
        new_col = phone_col
        print(f"  output    = col {new_col}  ({header[new_col-1]!r}) [filling empty cells in existing column]")
    else:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value=args.phone_header)
        print(f"  output    = col {new_col}  ({args.phone_header!r}) [new column]")

    # Build candidate list (only rows that need enrichment AND have inputs)
    candidates = []
    for row_idx in range(2, ws.max_row + 1):
        existing_phone = ws.cell(row=row_idx, column=phone_col).value if phone_col else None
        if existing_phone and str(existing_phone).strip():
            continue
        email = ws.cell(row=row_idx, column=email_col).value
        if not email or "@" not in str(email):
            continue
        if name_col:
            full_name = str(ws.cell(row=row_idx, column=name_col).value or "")
            first, last = split_name(full_name)
        else:
            first = str(ws.cell(row=row_idx, column=first_col).value or "")
            last = str(ws.cell(row=row_idx, column=last_col).value or "")
        if not first and not last:
            continue
        company = (str(ws.cell(row=row_idx, column=company_col).value or "")
                   if company_col else "")
        domain = email_to_domain(str(email))
        candidates.append({
            "row_idx": row_idx,
            "first_name": first,
            "last_name": last,
            "company": company or domain or "",
            "company_domain": domain or "",
        })

    if args.limit:
        candidates = candidates[: args.limit]
    print(f"\n{len(candidates)} rows enrichable (after dedup + input filters)")

    if args.dry_run:
        print("\n[dry-run] showing first 5 candidate rows:")
        for c in candidates[:5]:
            print(f"  row {c['row_idx']:>4}: {c['first_name']} {c['last_name']} | "
                  f"company={c['company']!r} domain={c['company_domain']!r}")
        return

    if not candidates:
        wb.save(args.output)
        print(f"Nothing to enrich. Output saved (header-only column added): {args.output}")
        return

    # Worst-case budget check before submitting:
    # 10 credits per found phone × candidate count.
    if args.budget_credits is not None:
        worst_case = len(candidates) * 10
        if worst_case > args.budget_credits:
            allowed = args.budget_credits // 10
            print(f"\n!!! Budget {args.budget_credits} credits → can fund at most "
                  f"{allowed} phone hits. Truncating to first {allowed} candidates.")
            candidates = candidates[:allowed]

    # Submit in batches of up to 100
    print(f"\nSubmitting {len(candidates)} candidates in batches of {BC_BATCH_SIZE}...")
    by_row = {c["row_idx"]: c for c in candidates}
    total_credits = 0.0
    total_found = 0

    for i in range(0, len(candidates), BC_BATCH_SIZE):
        batch = candidates[i:i + BC_BATCH_SIZE]
        payload = {
            "data": [
                {
                    "first_name": c["first_name"],
                    "last_name": c["last_name"],
                    "company": c["company"],
                    "company_domain": c["company_domain"],
                    "custom_fields": {"row_idx": str(c["row_idx"])},
                }
                for c in batch
            ],
            "enrich_email_address": False,
            "enrich_phone_number": True,
        }
        try:
            rid = bc_submit(payload, api_key)
            print(f"  batch {i//BC_BATCH_SIZE + 1}: submitted {len(batch)} leads (rid={rid})")
            result = bc_poll(rid, api_key)
        except Exception as e:
            print(f"  ! batch {i//BC_BATCH_SIZE + 1} failed: {e}", file=sys.stderr)
            continue

        cc = float(result.get("credits_consumed") or 0)
        total_credits += cc
        # BC's response shape: items in `data`, custom_fields is a LIST of
        # `{name, value, position}` dicts (not a flat dict like I first wrote).
        for ld in result.get("data") or []:
            row_idx = None
            for cf in ld.get("custom_fields") or []:
                if isinstance(cf, dict) and cf.get("name") == "row_idx":
                    try:
                        row_idx = int(cf.get("value"))
                        break
                    except (TypeError, ValueError):
                        pass
            if row_idx is None:
                continue
            phone = ld.get("contact_phone_number") or ""
            if phone:
                ws.cell(row=row_idx, column=new_col, value=str(phone))
                total_found += 1
        # Save incrementally so progress isn't lost on crash
        wb.save(args.output)
        print(f"  batch {i//BC_BATCH_SIZE + 1}: +{cc:.0f} credits, total_found={total_found}, saved.")

    print(f"\n=== summary ===")
    print(f"  candidates submitted: {len(candidates)}")
    print(f"  phones found:         {total_found}")
    print(f"  credits spent:        {total_credits:.1f}")
    print(f"  output:               {args.output}")


if __name__ == "__main__":
    main()
