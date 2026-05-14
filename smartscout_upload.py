"""Upsert SmartScout Amazon brand market data into smartscout_brands.

Re-runnable: matches on brand_norm (normalized brand name). Existing rows have
every market column overwritten and last_seen_at + updated_at bumped. Brands
that disappear from a new export are kept (stale data); their last_seen_at
just stops advancing.

CLI: python run.py upload-smartscout <file>
Accepts .csv or .xlsx.
"""

from __future__ import annotations

import csv
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from supabase import create_client


HEADER_MAP: dict[str, str] = {
    "Brand": "brand_original",
    "Primary Category": "primary_category",
    "Primary Subcategory": "primary_subcategory",
    "Amazon In-Stock Rate": "amazon_in_stock_rate",
    "Average Number of Sellers": "average_number_of_sellers",
    "Average Price": "average_price",
    "Estimated Monthly Revenue": "estimated_monthly_revenue",
    "Estimated Monthly Units Sold": "estimated_monthly_units_sold",
    "1 Month Growth": "one_month_growth",
    "12 Month Growth": "twelve_month_growth",
    "Trailing 12 Months": "trailing_12_months",
    "Average Package Volume": "average_package_volume",
    "Average Rating": "average_rating",
    "Total Ratings Count": "total_ratings_count",
    "Average Number of FBA Sellers": "average_number_of_fba_sellers",
    "Total Product Count": "total_product_count",
    "Brand Score": "brand_score",
    "Storefront": "storefront",
    "Dominant Seller": "dominant_seller",
    "Dominant Seller Sales Percentage": "dominant_seller_sales_percentage",
    "Dominant Seller Country": "dominant_seller_country",
}

NUMERIC_COLS: set[str] = {
    "amazon_in_stock_rate",
    "average_number_of_sellers",
    "average_price",
    "estimated_monthly_revenue",
    "estimated_monthly_units_sold",
    "one_month_growth",
    "twelve_month_growth",
    "trailing_12_months",
    "average_package_volume",
    "average_rating",
    "average_number_of_fba_sellers",
    "brand_score",
    "dominant_seller_sales_percentage",
}
INTEGER_COLS: set[str] = {"total_ratings_count", "total_product_count"}

CHUNK_SIZE = 500
MAX_RETRIES = 5
RETRY_BASE_DELAY = 2.0


_SUFFIX_RE = re.compile(
    r"\b(inc|incorporated|llc|ltd|limited|corp|corporation|co|company|"
    r"brand|brands|group|holdings|the)\b",
    re.IGNORECASE,
)


def normalize_brand(s: str | None) -> str:
    """Normalize a brand string for matching. Used at upload AND at resolve."""
    if not s:
        return ""
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = _SUFFIX_RE.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_numeric(raw: str | None) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() in {"n/a", "na", "-", "—", "null", "none"}:
        return None
    s = s.replace("$", "").replace(",", "").replace("%", "").strip()
    if s in {"", "-", "."}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_integer(raw: str | None) -> int | None:
    f = parse_numeric(raw)
    if f is None:
        return None
    return int(f)


def upsert_with_retry(supabase, chunk: list[dict]) -> None:
    last_exc: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            supabase.table("smartscout_brands").upsert(
                chunk, on_conflict="brand_norm"
            ).execute()
            return
        except Exception as exc:
            last_exc = exc
            if attempt == MAX_RETRIES:
                break
            delay = RETRY_BASE_DELAY * (2 ** (attempt - 1))
            print(f"    retry {attempt}/{MAX_RETRIES - 1} after error: "
                  f"{type(exc).__name__}: {exc} (sleeping {delay:.1f}s)")
            time.sleep(delay)
    raise last_exc  # type: ignore[misc]


def _resolve_path(raw: str) -> Path:
    p = Path(raw)
    if p.exists():
        return p
    for ext in (".csv", ".xlsx"):
        candidate = Path(f"{raw}{ext}")
        if candidate.exists():
            return candidate
    sys.exit(f"File not found: {raw} (also tried .csv, .xlsx)")


def _clean_header(h) -> str:
    return str(h).strip() if h is not None else ""


def _find_header_row(rows_iter, max_scan: int = 5) -> list[str]:
    """Scan up to max_scan rows for one that contains a 'Brand' column.
    SmartScout exports prepend a 'DATABASE / SMARTSCOUT' banner row."""
    for _ in range(max_scan):
        try:
            row = next(rows_iter)
        except StopIteration:
            break
        cleaned = [_clean_header(h) for h in row]
        if "Brand" in cleaned:
            return cleaned
    sys.exit("Could not find header row containing 'Brand' in first 5 rows.")


def _iter_csv_rows(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = _find_header_row(reader)
        for row in reader:
            yield {h: (v if v is not None else "") for h, v in zip(headers, row)}


def _iter_xlsx_rows(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = _find_header_row(rows)
    for row in rows:
        yield {h: ("" if v is None else str(v)) for h, v in zip(headers, row)}


def iter_rows(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _iter_csv_rows(path)
    if suffix == ".xlsx":
        return _iter_xlsx_rows(path)
    sys.exit(f"Unsupported file type: {suffix} (expected .csv or .xlsx)")


def to_db_row(src: dict, now_iso: str) -> dict | None:
    brand_raw = (src.get("Brand") or "").strip()
    if not brand_raw:
        return None
    brand_norm = normalize_brand(brand_raw)
    if not brand_norm:
        return None

    out: dict = {
        "brand_norm": brand_norm,
        "brand_original": brand_raw,
        "last_seen_at": now_iso,
        "updated_at": now_iso,
    }
    for header, col in HEADER_MAP.items():
        if col == "brand_original":
            continue
        val = src.get(header)
        if col in NUMERIC_COLS:
            out[col] = parse_numeric(val)
        elif col in INTEGER_COLS:
            out[col] = parse_integer(val)
        else:
            v = (str(val).strip() if val is not None else "")
            out[col] = v if v else None
    return out


def chunked(items: list, size: int):
    for i in range(0, len(items), size):
        yield items[i : i + size]


def main(file_arg: str) -> None:
    load_dotenv()
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not url or not key:
        sys.exit("SUPABASE_URL and SUPABASE_KEY must be set in .env")

    path = _resolve_path(file_arg)
    print(f"Reading {path}")

    now_iso = datetime.now(timezone.utc).isoformat()

    rows_by_brand: dict[str, dict] = {}
    skipped = 0
    duplicates_in_file = 0

    for src in iter_rows(path):
        db_row = to_db_row(src, now_iso)
        if db_row is None:
            skipped += 1
            continue
        bn = db_row["brand_norm"]
        if bn in rows_by_brand:
            duplicates_in_file += 1
        rows_by_brand[bn] = db_row  # last wins

    rows = list(rows_by_brand.values())
    total = len(rows)
    print(f"Parsed: {total} unique brands, {skipped} skipped (no Brand), "
          f"{duplicates_in_file} duplicate brand_norm collisions (last wins)")

    if not rows:
        print("Nothing to upload.")
        return

    supabase = create_client(url, key)
    upserted = 0
    for i, chunk in enumerate(chunked(rows, CHUNK_SIZE), start=1):
        upsert_with_retry(supabase, chunk)
        upserted += len(chunk)
        print(f"  chunk {i}: {upserted}/{total} upserted")

    print(f"Done. {upserted} rows upserted into smartscout_brands.")
    print("Run `python run.py resolve-smartscout` next to match leads to brands.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("Usage: python smartscout_upload.py <file.csv|.xlsx>")
    main(sys.argv[1])
