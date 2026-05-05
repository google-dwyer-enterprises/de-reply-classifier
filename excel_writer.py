"""Phase 6: Excel export — writeback and fresh modes."""

from __future__ import annotations

import os
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from supabase import create_client

from classify import clean_body
from config import is_excluded_sender

EXCLUDED_STATUS_LABELS = {"oof", "customer_service"}

STATUS_PRIORITY = [
    "booked", "interested", "interested_past", "not_now",
    "wrong_person", "no_longer_there", "not_interested",
    "unsubscribe", "other", "oof", "customer_service",
]
STATUS_RANK = {label: i for i, label in enumerate(STATUS_PRIORITY)}

FRESH_COLUMNS = [
    "lead_email", "clients", "campaigns",
    "status1", "status2", "status3", "status4",
    "status_confidence", "score",
    "last_reply_date", "total_replies",
    "thread_id", "subject", "body_preview",
]

WRITEBACK_COLUMNS = [
    "status1", "status2", "status3", "status4",
    "status_confidence", "score", "reason", "clients", "campaigns",
    "last_reply_date", "total_replies", "thread_id",
]

EMAIL_HEADER_NAMES = {"emails", "email", "email address", "lead_email", "email_address"}


# --------------------------------------------------------------------------- #
# Supabase + per-lead summary
# --------------------------------------------------------------------------- #

def get_supabase():
    load_dotenv()
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_KEY", "").strip()
    if not url or not key:
        sys.exit("FATAL: SUPABASE_URL and SUPABASE_KEY must be set in .env")
    return create_client(url, key)


def _paginate_all(query_builder, page_size: int = 1000, label: str = "rows") -> list[dict]:
    out: list[dict] = []
    start = 0
    while True:
        resp = query_builder.range(start, start + page_size - 1).execute()
        batch = resp.data or []
        out.extend(batch)
        print(f"  fetched {len(out):>6} {label}...", flush=True)
        if len(batch) < page_size:
            return out
        start += page_size


def fetch_per_lead_summary(supabase) -> dict[str, dict]:
    print("Fetching label_scores from Supabase...", flush=True)
    score_rows = supabase.table("label_scores").select("label,score").execute().data or []
    score_map = {r["label"]: r["score"] for r in score_rows}
    if not score_map:
        sys.exit("FATAL: label_scores table is empty. Run `python scripts/seed_label_scores.py`.")

    print("Fetching replies from Supabase...", flush=True)
    replies = _paginate_all(supabase.table("replies").select("*"), label="replies")
    print("Fetching classifications from Supabase...", flush=True)
    classifications = _paginate_all(
        supabase.table("classifications")
        .select("reply_id, label, confidence, reason, classified_at")
        .order("classified_at", desc=False),
        label="classifications",
    )
    print(f"Building per-lead summary from {len(replies)} replies + {len(classifications)} classifications...", flush=True)
    # Iterating ascending by classified_at means later writes overwrite earlier
    # ones in the dict — so the newest classification per reply_id wins. This
    # lets v2 and v3 (or future versions) coexist in the table without ambiguity.
    class_by_reply = {c["reply_id"]: c for c in classifications}

    by_email: dict[str, list] = defaultdict(list)
    excluded_leads: set[str] = set()
    excluded_replies = 0
    for r in replies:
        email = (r.get("lead_email") or "").strip().lower()
        if not email:
            continue
        if is_excluded_sender(email):
            excluded_leads.add(email)
            excluded_replies += 1
            continue
        c = class_by_reply.get(r["id"])
        r["_label"] = c["label"] if c else None
        r["_confidence"] = c["confidence"] if c else None
        r["_reason"] = c.get("reason") if c else None
        r["_score"] = score_map.get(r["_label"] or "", 0)
        by_email[email].append(r)

    if excluded_leads:
        print(f"Excluded {len(excluded_leads)} bot/internal senders ({excluded_replies} replies). Examples: {sorted(excluded_leads)[:5]}", flush=True)

    summary: dict[str, dict] = {}
    for email, rlist in by_email.items():
        rlist.sort(key=lambda x: x.get("reply_timestamp") or "", reverse=True)

        # Top-3 distinct classifier labels by STATUS_RANK (booked first).
        # Tie-breaker within same label: most recent reply_timestamp.
        seen_labels: dict[str, dict] = {}
        for r in rlist:
            lbl = r.get("_label")
            if not lbl:
                continue
            if lbl not in seen_labels:
                seen_labels[lbl] = r  # rlist is newest-first so this is the most recent
        ranked = sorted(seen_labels.items(), key=lambda kv: STATUS_RANK.get(kv[0], 999))
        top3 = ranked[:3] + [(None, None)] * (3 - len(ranked))
        status1, status2, status3 = (t[0] for t in top3)
        picked = top3[0][1]  # reply that produced status1

        # status4 = latest non-null Instantly lead_status across this lead's replies
        status4 = next(
            (r.get("lead_status") for r in rlist if r.get("lead_status")),
            None,
        )

        if picked is not None:
            confidence = float(picked["_confidence"]) if picked.get("_confidence") is not None else None
            reason = picked.get("_reason") or ""
            last_reply_date = picked.get("reply_timestamp")
            thread_id = picked.get("thread_id") or ""
            subject = picked.get("subject") or ""
            body = picked.get("body") or ""
        else:
            # No classifier label on any reply — fall back to most recent reply for metadata.
            picked = rlist[0]
            confidence = None
            reason = ""
            last_reply_date = picked.get("reply_timestamp")
            thread_id = picked.get("thread_id") or ""
            subject = picked.get("subject") or ""
            body = picked.get("body") or ""

        preview = clean_body(body)[:200]

        by_thread: dict[str, int] = {}
        for r in rlist:
            key = r.get("thread_id") or f"__noth_{r['id']}"
            s = r.get("_score", 0)
            if key not in by_thread or s > by_thread[key]:
                by_thread[key] = s
        score = sum(by_thread.values())

        clients_set = sorted({(r.get("client") or "").strip() for r in rlist if (r.get("client") or "").strip()})
        campaigns_set = sorted({(r.get("campaign_name") or "").strip() for r in rlist if (r.get("campaign_name") or "").strip()})

        summary[email] = {
            "lead_email": email,
            "clients": "; ".join(clients_set),
            "campaigns": "; ".join(campaigns_set),
            "status1": status1,
            "status2": status2,
            "status3": status3,
            "status4": status4,
            "status_confidence": confidence,
            "score": score,
            "reason": reason,
            "last_reply_date": last_reply_date,
            "total_replies": len(rlist),
            "thread_id": thread_id,
            "subject": subject,
            "body_preview": preview,
        }
    return summary


# --------------------------------------------------------------------------- #
# Cell coercion + formatting
# --------------------------------------------------------------------------- #

def _coerce(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (datetime, date)):
        if isinstance(v, datetime) and v.tzinfo is not None:
            v = v.replace(tzinfo=None)
        return v
    s = str(v).strip()
    if "T" in s and "-" in s:
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            if dt.tzinfo is not None:
                dt = dt.replace(tzinfo=None)
            return dt
        except Exception:
            pass
    s = "".join(c for c in s if ord(c) >= 0x20 or c in "\n\t")
    return s


def _auto_width(ws, columns: list[str]) -> None:
    for idx, _ in enumerate(columns, 1):
        col_letter = get_column_letter(idx)
        max_len = len(columns[idx - 1])
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            l = len(str(cell.value))
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(max_len + 2, 80)


# --------------------------------------------------------------------------- #
# Fresh export
# --------------------------------------------------------------------------- #

def export_fresh(output_path: str) -> None:
    supabase = get_supabase()
    summary = fetch_per_lead_summary(supabase)
    rows = list(summary.values())

    def sort_key(r):
        rank = STATUS_RANK.get(r.get("status1") or "", 999)
        lr = r.get("last_reply_date")
        ts = -_iso_to_ts(lr)
        return (rank, ts)

    rows.sort(key=sort_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Classified Leads"
    ws.append(FRESH_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    print(f"Writing {len(rows)} rows to workbook...", flush=True)
    for r in rows:
        ws.append([_coerce(r.get(c)) for c in FRESH_COLUMNS])

    ws.freeze_panes = "A2"
    _auto_width(ws, FRESH_COLUMNS)

    out = Path(output_path)
    out.parent.mkdir(exist_ok=True, parents=True)
    wb.save(out)

    label_counts = Counter((r.get("status1") or "") for r in rows)
    print()
    print("=" * 60)
    print("FRESH EXPORT SUMMARY")
    print("=" * 60)
    print(f"Output: {out}")
    print(f"Total rows exported: {len(rows)}")
    print("Label breakdown:")
    for lbl, n in label_counts.most_common():
        print(f"  {lbl:20} {n}")

    def _print_row(r):
        conf = r["status_confidence"]
        conf_s = f"{conf:.2f}" if conf is not None else ""
        date_s = str(r["last_reply_date"])[:19] if r["last_reply_date"] else ""
        s4 = (r.get("status4") or "")[:20]
        print(
            f"  {r['lead_email']:35.35} | {(r.get('status1') or ''):17.17} | {conf_s:>4} | "
            f"s4={s4:20.20} | {date_s} | subject={r['subject'][:40]!r}"
        )

    print()
    print("First 5 rows (highest-priority after sort):")
    print("-" * 60)
    for r in rows[:5]:
        _print_row(r)
    print()
    print("Last 5 rows (lowest-priority after sort):")
    print("-" * 60)
    for r in rows[-5:]:
        _print_row(r)


def _iso_to_ts(v) -> float:
    if not v:
        return 0.0
    try:
        s = str(v).replace("Z", "+00:00")
        return datetime.fromisoformat(s).timestamp()
    except Exception:
        return 0.0


# --------------------------------------------------------------------------- #
# Writeback export
# --------------------------------------------------------------------------- #

def export_writeback(input_path: str, tab: str, header_row: int, output_path: str | None = None) -> None:
    supabase = get_supabase()
    summary = fetch_per_lead_summary(supabase)

    in_path = Path(input_path)
    if not in_path.is_file():
        sys.exit(f"FATAL: input file not found: {in_path}")

    wb = load_workbook(in_path)
    if tab not in wb.sheetnames:
        sys.exit(f"FATAL: tab {tab!r} not found. Available: {wb.sheetnames}")
    ws = wb[tab]

    header_cells = list(ws[header_row])
    email_col_idx = None
    for idx, cell in enumerate(header_cells):
        h = str(cell.value or "").strip().lower()
        if h in EMAIL_HEADER_NAMES:
            email_col_idx = idx
            break
    if email_col_idx is None:
        sys.exit(
            f"FATAL: no email column in row {header_row}. "
            f"Headers: {[c.value for c in header_cells[:20]]}"
        )

    header_map = {
        str(c.value or "").strip().lower(): i for i, c in enumerate(header_cells)
    }
    target_col_idx: dict[str, int] = {}
    last_col_idx = len(header_cells)
    for col_name in WRITEBACK_COLUMNS:
        if col_name.lower() in header_map:
            target_col_idx[col_name] = header_map[col_name.lower()]
        else:
            target_col_idx[col_name] = last_col_idx
            cell = ws.cell(row=header_row, column=last_col_idx + 1, value=col_name)
            cell.font = Font(bold=True)
            last_col_idx += 1

    matched = 0
    no_reply = 0
    empty = 0
    label_counts: Counter = Counter()
    matched_samples: list[tuple[int, dict]] = []

    import time
    total_rows = ws.max_row - header_row
    start_ts = time.monotonic()
    print(f"Scanning {total_rows} data rows in tab {tab!r}...", flush=True)
    progress_every = max(500, total_rows // 40)

    for excel_row in range(header_row + 1, ws.max_row + 1):
        email_cell = ws.cell(row=excel_row, column=email_col_idx + 1).value
        if not email_cell:
            empty += 1
            continue
        email_norm = str(email_cell).strip().lower()
        lead = summary.get(email_norm)

        if lead:
            matched += 1
            label_counts[lead.get("status1") or ""] += 1
            for col_name in WRITEBACK_COLUMNS:
                ws.cell(
                    row=excel_row,
                    column=target_col_idx[col_name] + 1,
                    value=_coerce(lead.get(col_name)),
                )
            if len(matched_samples) < 5:
                matched_samples.append((excel_row, lead))
        else:
            no_reply += 1
            ws.cell(
                row=excel_row,
                column=target_col_idx["status1"] + 1,
                value="no_reply",
            )

        processed = excel_row - header_row
        if processed % progress_every == 0 or processed == total_rows:
            elapsed = time.monotonic() - start_ts
            rate = processed / elapsed if elapsed > 0 else 0
            remaining = (total_rows - processed) / rate if rate > 0 else 0
            pct = processed / total_rows * 100 if total_rows else 100
            print(
                f"  row {processed:>6}/{total_rows}  ({pct:5.1f}%)  "
                f"matched={matched}  rate={rate:,.0f} rows/s  ETA={remaining:,.0f}s",
                flush=True,
            )

    print(f"Saving workbook...", flush=True)

    if output_path:
        out = Path(output_path)
    else:
        safe_tab = "".join(c if c.isalnum() or c in "-_" else "_" for c in tab).strip("_") or "tab"
        out = (
            in_path.parent
            / f"{in_path.stem}__classified_{safe_tab}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
    wb.save(out)

    total = matched + no_reply + empty
    denom = matched + no_reply
    pct = (matched / denom * 100) if denom else 0.0

    print()
    print("=" * 60)
    print("WRITEBACK SUMMARY")
    print("=" * 60)
    print(f"Input:  {in_path}")
    print(f"Output: {out}")
    print(f"Tab:    {tab}  (header row {header_row})")
    print(f"Total rows scanned: {total}")
    print(f"  matched (status written):     {matched}")
    print(f"  no_reply (not in classifications): {no_reply}")
    print(f"  skipped (empty email cell):   {empty}")
    print(f"  match rate: {pct:.1f}%")
    if label_counts:
        print("Label breakdown (matched rows):")
        for lbl, n in label_counts.most_common():
            print(f"  {lbl:20} {n}")

    if matched_samples:
        print()
        print("First 5 matched rows:")
        print("-" * 60)
        for row_num, lead in matched_samples:
            conf = lead["status_confidence"]
            conf_s = f"{conf:.2f}" if conf is not None else ""
            date_s = str(lead["last_reply_date"])[:19] if lead["last_reply_date"] else ""
            s4 = (lead.get("status4") or "")[:20]
            print(
                f"  row={row_num:>5} | {lead['lead_email']:35.35} | {(lead.get('status1') or ''):17.17} | "
                f"{conf_s:>4} | s4={s4:20.20} | {date_s}"
            )
